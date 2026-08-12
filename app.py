import csv
import io
import json
import os
import secrets
import uuid
import zipfile
import bcrypt
import subprocess
from datetime import datetime
from functools import wraps
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, abort, Response
)
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)

_secret = os.environ.get("SECRET_KEY", "")
if not _secret or _secret == "change-me-in-production":
    import sys
    print("FEHLER: SECRET_KEY nicht gesetzt oder Standard-Wert. Bitte in .env setzen.", file=sys.stderr)
    if os.environ.get("FLASK_ENV") != "development":
        sys.exit(1)
app.secret_key = _secret or "dev-only-insecure-key"

app.config.update(
    SESSION_COOKIE_HTTPONLY  = True,
    SESSION_COOKIE_SAMESITE  = "Lax",
    SESSION_COOKIE_SECURE    = os.environ.get("HTTPS", "1") == "1",
)

DATA_DIR      = os.path.join(os.path.dirname(__file__), "data")
SIGNALS_FILE  = os.path.join(DATA_DIR, "signals.json")
USERS_FILE    = os.path.join(DATA_DIR, "users.json")
ENTITIES_FILE = os.path.join(DATA_DIR, "entities.json")
SECTIONS_FILE = os.path.join(DATA_DIR, "sections.seed.json")  # sections are static
TODOS_FILE    = os.path.join(DATA_DIR, "todos.json")
SETTINGS_FILE = os.path.join(DATA_DIR, "settings.json")
SOURCES_FILE             = os.path.join(DATA_DIR, "sources.json")
AGENT_CONFIGS_FILE       = os.path.join(DATA_DIR, "agent_configs.json")
SIGNAL_MATCHES_FILE      = os.path.join(DATA_DIR, "signal_matches.json")
SIGNAL_FINAL_FILE        = os.path.join(DATA_DIR, "signal_final.json")
SIGNAL_FINAL_ARCHIVE_DIR = os.path.join(DATA_DIR, "signal_final_archive")
AGENT_REPORTS_FILE       = os.path.join(DATA_DIR, "agent_reports.json")

DEFAULT_TODO_CATEGORIES = [
    "Abrechnung & Vergütung",
    "Compliance & Regulatorik",
    "Praxisorganisation",
    "IT & Digitalisierung",
    "Personal & Fortbildung",
]

def _init_data():
    """Copy seed files to data files on first run if data files don't exist."""
    import shutil
    for name in ("signals", "users", "entities", "todos", "sources"):
        target = os.path.join(DATA_DIR, f"{name}.json")
        seed   = os.path.join(DATA_DIR, f"{name}.seed.json")
        if not os.path.exists(target) and os.path.exists(seed):
            shutil.copy2(seed, target)
    # settings is a dict, not a list
    if not os.path.exists(SETTINGS_FILE):
        seed = os.path.join(DATA_DIR, "settings.seed.json")
        if os.path.exists(seed):
            shutil.copy2(seed, SETTINGS_FILE)
        else:
            _save(SETTINGS_FILE, {"default_todo_categories": DEFAULT_TODO_CATEGORIES})

_init_data()

DEPLOY_TOKEN = os.environ.get("DEPLOY_TOKEN", "")

STATUS_ORDER = ["action", "announced", "develop", "radar", "active"]
CATEGORIES = {
    "krankenkassen": "Krankenkassen & GKV",
    "digital":       "Digitalisierung & TI",
    "gesetze":       "Gesetze & Regulatorien",
    "personal":      "Personal & Tarife",
    "praxis":        "Praxismanagement",
}
STATUS_LABELS = {
    "action":    "Handlungsbedarf",
    "announced": "Angekündigt",
    "develop":   "In Entwicklung",
    "radar":     "Im Radar",
    "active":    "Verfügbar / In Kraft",
}
SIGNAL_PRIORITY_LABELS = {
    "muss":   "Muss",
    "sollte": "Sollte",
    "kann":   "Kann",
}
ENTWICKLUNGSSTAND_LABELS = {
    "beobachtung":      "Beobachtung",
    "entwurf":          "Entwurf",
    "konsultation":     "Konsultation",
    "beschlossen":      "Beschlossen",
    "veroeffentlicht":  "Verkündet / Veröffentlicht",
    "in_kraft_kuenftig":"In Kraft ab [Datum]",
    "in_kraft":         "In Kraft",
    "aufgehoben":       "Aufgehoben / Abgelöst",
}
HANDLUNGSZEITPUNKT_LABELS = {
    "jetzt":       "Jetzt handeln",
    "vorbereiten": "Bald handeln",
    "beobachten":  "Beobachten",
    "keine_aktion":"Keine Aktion",
}
AUFWAND_LABELS = {
    "unter_15":        "< 15 Min.",
    "15_60":           "15–60 Min.",
    "mehrere_stunden": "Mehrere Stunden",
}
SIGNAL_ROLLEN = [
    "Praxisleitung",
    "Ärzt:innen",
    "MFA / ZFA",
    "Abrechnung",
    "Praxismanagement",
    "Datenschutz",
    "QM",
]
ROLE_LABELS = {
    "superadmin": "Superadmin",
    "admin":      "Admin",
    "user":       "Nutzer",
}


# ── Data helpers ──────────────────────────────────────────────────────────────

def _load(path):
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _save(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_signals():   return _load(SIGNALS_FILE)
def save_signals(d):  _save(SIGNALS_FILE, d)
def load_users():     return _load(USERS_FILE)
def save_users(d):    _save(USERS_FILE, d)
def load_entities():  return _load(ENTITIES_FILE)
def save_entities(d): _save(ENTITIES_FILE, d)
def load_sections():  return _load(SECTIONS_FILE)
def load_todos():     return _load(TODOS_FILE)
def save_todos(d):    _save(TODOS_FILE, d)
def load_sources():   return _load(SOURCES_FILE)
def save_sources(d):  _save(SOURCES_FILE, d)

_SEARCH_OUTPUT_FORMAT = (
    "## Ausgabeformat\n"
    "Für jedes gefundene Signal erstelle einen Eintrag mit folgenden Feldern:\n\n"
    "**Titel:** [Sprechender Action-Titel im Stil eines Managementpräsentations-Titels – prägnant, aktiv, handlungsorientiert, 8–12 Wörter]\n"
    "**Zusammenfassung:** [500–1000 Zeichen. Was passiert konkret, was bedeutet das für Gesundheitseinrichtungen?]\n"
    "**Veröffentlicht am:** [TT.MM.JJJJ – Datum der Veröffentlichung durch die Quelle; immer befüllen wenn bekannt]\n"
    "**Beschlossen am:** [TT.MM.JJJJ – Datum des Beschlusses/der Entscheidung, z.B. G-BA-Beschlussdatum; leer lassen wenn nicht zutreffend]\n"
    "**Im Bundesanzeiger:** [TT.MM.JJJJ – Datum der amtlichen Bekanntmachung im Bundesanzeiger oder EU-Amtsblatt; nur für Gesetze/Verordnungen]\n"
    "**Gilt ab:** [TT.MM.JJJJ – Datum des Inkrafttretens oder der Wirksamkeit; z.B. 01.10.2026 für zukünftige Regelungen]\n"
    "**Frist bis:** [TT.MM.JJJJ – Handlungsfrist oder Stellungnahmefrist; z.B. bei AWMF-Konsultationen oder Übergangszeiträumen]\n"
    "**Entwicklungsstand:** [Genau eines von: Beobachtung | Entwurf | Konsultation | Beschlossen | Verkündet / Veröffentlicht | In Kraft ab [Datum] | In Kraft | Aufgehoben / Abgelöst]\n"
    "  → Beobachtung: früher Hinweis, noch nicht formalisiert\n"
    "  → Entwurf: Referentenentwurf oder Beschlussentwurf veröffentlicht\n"
    "  → Konsultation: formale Konsultationsphase mit Frist (z.B. AWMF, EU)\n"
    "  → Beschlossen: formal beschlossen (G-BA, Bundestag), noch nicht im Bundesanzeiger\n"
    "  → Verkündet / Veröffentlicht: im Bundesanzeiger / Amtsblatt, noch nicht wirksam oder Sicherheitsinformation\n"
    "  → In Kraft ab [Datum]: verkündet, zukünftiges Inkrafttreten bekannt\n"
    "  → In Kraft: unmittelbar rechtlich wirksam\n"
    "  → Aufgehoben / Abgelöst: widerrufen oder durch neue Regelung ersetzt\n"
    "**Handlungszeitpunkt:** [Genau eines von: Sofort | Kurzfristig (< 3 Monate) | Mittelfristig (3–12 Monate) | Langfristig (> 12 Monate) | Beobachten]\n"
    "**Quelle 1:** [Name der Quelle, z.B. G-BA, BMG, gematik, BZÄK]\n"
    "**Quellenlink 1:** [Direkte URL zum Originaldokument oder zur Meldung]\n"
    "**Agent:** [Exakter Name dieses Agenten: scrape | feed | pdf | search]\n"
    "**Region:** [Geltungsbereich: Bundesweit | Bayern | Nordrhein-Westfalen | … | EU / International]\n\n"
    "Alle anderen Felder (Priorität, Kategorie, Fachbereich, Betroffene Rollen, Aufwand, Nächster Schritt) "
    "werden von den Such-Agenten nicht befüllt und bleiben leer.\n\n"
    "**Granularität — Wann aufteilen, wann zusammenfassen:**\n"
    "Erstelle **ein Signal pro eigenständiger Maßnahme**. Trenne immer dann, wenn sich mindestens eines dieser Merkmale unterscheidet:\n"
    "– Empfänger (z.B. nur Zahnärzte vs. nur Hausärzte)\n"
    "– Handlung (z.B. Abrechnungsänderung vs. Gerätepflicht vs. Dokumentationspflicht)\n"
    "– Frist oder Inkrafttreten (z.B. sofort vs. ab 01.01.2027)\n"
    "– Rechtlicher Rahmen (z.B. Gesetz vs. G-BA-Richtlinie vs. Sicherheitsmitteilung)\n\n"
    "**Beispiel Spargesetz:** Ein Artikel beschreibt Arzneimittelausschlüsse, Cannabis-Neuregelungen und Wegfall von Hygienezuschlägen. "
    "Das sind drei separate Signale mit unterschiedlichen Empfängern und Fristen — nicht ein Signal 'Spargesetz'.\n\n"
    "**Beispiel G-BA-Beschlüsse:** Eine Übersichtsseite listet sieben Arzneimittelbewertungen. "
    "Jede Bewertung ist ein eigenständiges Signal — auch wenn du die Zielgruppenrelevanz niedrig einschätzt. "
    "**Deine Aufgabe ist sammeln, nicht filtern.** Relevanzentscheidungen trifft der nachgelagerte Bewertungs-Agent.\n\n"
    "**Zusammenfassen** nur wenn: dieselbe Maßnahme, dieselben Empfänger, dasselbe Inkrafttreten — "
    "aber von mehreren Quellen berichtet. In diesem Fall eine Quelle als Primärquelle, weitere als Quelle 2 / Quelle 3.\n\n"
    "---\n\n"
    "## Crawl-Report\n"
    "Nach allen Signalen: Erstelle eine separate Tabelle im TSV-Format (Tab-getrennt) mit genau diesen 5 Spalten:\n"
    "Quelle\tAnzahl Signale\tHinweise/Probleme\tTimestamp\tAgent\n\n"
    "Führe **jede geprüfte Quelle** auf – auch solche ohne neue Signale oder mit technischen Problemen.\n"
    "– Anzahl Signale: Anzahl der in diesem Lauf für diese Quelle erstellten Signale (0 wenn keine)\n"
    "– Hinweise/Probleme: z.B. 'Seite nicht erreichbar', 'Login erforderlich', 'Kein neuer Inhalt', leer wenn ok\n"
    "– Timestamp: aktuelles Datum/Uhrzeit im Format JJMMDDhhmm (z.B. 2608061430)\n"
    "– Agent: exakter Name dieses Agenten (scrape | feed | pdf | search)"
)

_SEARCH_PERSONA_BASE = (
    "Du bist ein spezialisierter Such-Agent für den Branchenradar Gesundheitswesen. "
    "Deine Aufgabe ist es, relevante neue regulatorische Inhalte zu finden und als strukturierte Signale aufzubereiten. "
    "Ein Signal fasst ein regulatorisches Ereignis prägnant zusammen (500–1000 Zeichen) "
    "und gibt ihm einen sprechenden, handlungsorientierten Titel."
)

_SOURCE_MAINTENANCE_PERSONA = (
    "Du bist ein Quellen-Pflege-Agent für den Branchenradar Gesundheitswesen. "
    "Du erhältst eine Liste bestehender Quellen mit ihren aktuell hinterlegten Details. "
    "Deine Aufgabe ist es, jede Quelle zu analysieren und die hinterlegten Informationen zu vervollständigen, "
    "zu korrigieren und zu konkretisieren — insbesondere die URL-Endpunkte und Agenten-Hinweise."
)

_SOURCE_MAINTENANCE_METHOD = (
    "Gehe jede Quelle der übergebenen Liste durch:\n"
    "1. Prüfe ob die Quelle bereits Einträge in der Endpoints-Spalte hat. "
    "Quellen ohne Endpoints haben höchste Priorität — sie werden vom System aktuell gar nicht abgerufen.\n"
    "2. Suche nach konkreten Feed-/RSS-Endpunkten (typisch: /feed, /rss.xml, /atom.xml, Link-Tags im HTML-Header)\n"
    "3. Identifiziere relevante Unterseiten für Scrape-Endpunkte (Beschlüsse, Aktuelles, Pressemitteilungen, Rundschreiben, PDF-Verzeichnisse)\n"
    "4. Prüfe ob die zugewiesenen Agenten (scrape/feed/pdf/search) noch die richtigen sind\n"
    "5. Formuliere einen konkreten Agenten-Hinweis der beschreibt wo genau neue Inhalte zu finden sind\n"
    "6. Ergänze fehlende Felder soweit recherchierbar (Zugang, Aktualisierungsfrequenz, Hinweise)\n"
    "Hinweis: CSS-Selektoren für Scrape-Endpunkte werden automatisch durch Sol (Selektor-Finder) ermittelt — "
    "du musst diese nicht selbst recherchieren."
)

_SOURCE_MAINTENANCE_HINT = (
    "Priorität: Quellen ohne Endpoints-Einträge zuerst bearbeiten — diese werden aktuell nicht abgerufen. "
    "Fokussiere auf Vollständigkeit und Konkretheit: Eine URL wie 'https://www.g-ba.de' ist nutzlos, "
    "'https://www.g-ba.de/beschluesse/' ist brauchbar. "
    "Endpoint-Format: 'feed:https://example.com/rss.xml (RSS-Feed) | scrape:https://example.com/news/ (Aktuelles)'. "
    "Mehrere Endpunkte mit ' | ' trennen. CSS-Selektoren NICHT eintragen — das übernimmt Sol automatisch. "
    "Wenn du keinen direkten Feed findest, benenne die spezifische Unterseite die gecrawlt werden soll. "
    "Wenn eine Quelle weder Feed noch crawlbare Unterseite hat, empfehle 'search' als Agenten-Methode. "
    "Belasse Felder die du nicht mit Sicherheit ergänzen kannst leer — keine Vermutungen."
)

_SOURCE_TSV_HEADER = (
    "Kürzel\tName\tURL\tRegion\tEndpoints\tPrimärkategorie\tRelevante Rollen\tAgenten\t"
    "Priorität\tZugang\tAktualisierung\tStatus\tHinweise\t"
    "Agenten-Hinweis (scrape)\tAgenten-Hinweis (feed)\tAgenten-Hinweis (pdf)\tAgenten-Hinweis (search)\t"
    "Datum Hinzugefügt\tDatum Letzte Änderung\tKommentar"
)

_SOURCE_OUTPUT_RULES = (
    "Regeln:\n"
    "- 'URL': direkter Einstiegspunkt (Unterseite/Feed/Dokumentverzeichnis), nicht die Homepage\n"
    "- 'Feed-URL': RSS/Atom-URL falls vorhanden, sonst leer\n"
    "- 'Agenten': kommagetrennt aus scrape | feed | pdf | search\n"
    "- 'Relevante Rollen': kommagetrennt\n"
    "- 'Agenten-Hinweis (X)': nur für den/die zugewiesenen Agenten befüllen, konkret, max. 2 Sätze\n"
    "- 'Datum Hinzugefügt': leer lassen wenn Quelle bereits bestand — nur bei neuen Quellen heutiges Datum JJJJ-MM-TT\n"
    "- 'Datum Letzte Änderung': heutiges Datum JJJJ-MM-TT wenn du Felder angepasst hast, sonst leer\n"
    "- 'Kommentar': kurze Beschreibung was du geändert/hinzugefügt hast, z.B. 'Feed-URL ergänzt', 'Neu hinzugefügt'\n"
    "Liefere eine vollständige Liste aller Quellen (bestehende + neue/geänderte) als einzige TSV-Tabelle mit Kopfzeile."
)

_SOURCE_MAINTENANCE_OUTPUT = (
    "## Ausgabeformat\n"
    "Liefere eine vollständige gesamthafte Quellenliste als TSV (Tab-getrennt) mit genau dieser Kopfzeile:\n\n"
    + _SOURCE_TSV_HEADER + "\n\n"
    + _SOURCE_OUTPUT_RULES
)

_SOURCE_RESEARCH_PERSONA = (
    "Du bist ein Quellen-Recherche-Agent für den Branchenradar Gesundheitswesen. "
    "Deine Aufgabe ist es, neue relevante Quellen zu finden, die noch nicht in der bestehenden Quellenliste enthalten sind. "
    "Du kennst die deutschen Gesundheitseinrichtungen, ihre regulatorischen Stakeholder und die relevanten Informationskanäle."
)

_SOURCE_RESEARCH_METHOD = (
    "1. Analysiere den übergebenen Suchfokus (Thema, Einrichtungstyp, Region oder Behördentyp)\n"
    "2. Recherchiere gezielt nach Quellen die diesen Fokus abdecken und noch nicht in der Liste sind\n"
    "3. Für jede gefundene Quelle: prüfe ob sie regelmäßig relevante regulatorische Inhalte veröffentlicht\n"
    "4. Identifiziere sofort den konkreten Endpunkt (Feed-URL, Unterseite, Dokumentverzeichnis)\n"
    "5. Falls der Endpunkt ein Scrape-Endpunkt ist: Rufe die Seite ab und ermittle den CSS-Selektor "
    "der Artikel-Links (z.B. 'ul.news-list a'). Trage ihn direkt im Endpoints-Format ein.\n"
    "6. Bewerte Priorität: hoch (direkte Rechtswirkung), mittel (operative Relevanz), niedrig (Orientierung)\n"
    "7. Schließe Quellen aus die nur Werbung, rein redaktionelle Inhalte oder keine deutschen Bezüge haben"
)

_SOURCE_RESEARCH_HINT = (
    "Typische Quellen-Kategorien die oft fehlen: KV-Landesverbände, Landeszahnärztekammern, "
    "Landesgesundheitsministerien, Fachgesellschaften (z.B. DGAI, DGCH), Kassenzahnärztliche Vereinigungen, "
    "Bundesbehörden (BfArM, BZgA, PEI, DIMDI), Selbstverwaltungsorgane (G-BA, GKV-Spitzenverband). "
    "Prüfe auch: offizielle Bundesanzeiger-Rubriken, EUR-Lex für EU-Rechtsakte mit Deutschlandbezug, "
    "Krankenhausgesellschaften auf Landesebene. "
    "Liefere nur Quellen bei denen du den konkreten Endpunkt mit Sicherheit benennen kannst."
)

_SOURCE_RESEARCH_OUTPUT = (
    "## Ausgabeformat\n"
    "Liefere eine vollständige gesamthafte Quellenliste (Bestand + neue Quellen) als TSV (Tab-getrennt) mit genau dieser Kopfzeile:\n\n"
    + _SOURCE_TSV_HEADER + "\n\n"
    + _SOURCE_OUTPUT_RULES + "\n\n"
    "Zusatz für neue Quellen:\n"
    "- 'Kürzel': maximal 8 Zeichen, Großbuchstaben, eindeutig (z.B. BZÄK, KBVNL, GKVSV)\n"
    "- 'Datum Hinzugefügt': heutiges Datum JJJJ-MM-TT\n"
    "- 'Kommentar': 'Neu hinzugefügt'\n"
    "Bestehende Quellen aus der Bestandsliste vollständig übernehmen (keine Felder löschen)."
)

DEFAULT_AGENT_CONFIGS = {
    "output_format": _SEARCH_OUTPUT_FORMAT,
    "agents": {
        "source_maintenance": {
            "label": "Quellen-Pflege-Agent",
            "persona": _SOURCE_MAINTENANCE_PERSONA,
            "method":  _SOURCE_MAINTENANCE_METHOD,
            "hint":    _SOURCE_MAINTENANCE_HINT,
            "output_format": _SOURCE_MAINTENANCE_OUTPUT,
        },
        "source_research": {
            "label": "Quellen-Recherche-Agent",
            "persona": _SOURCE_RESEARCH_PERSONA,
            "method":  _SOURCE_RESEARCH_METHOD,
            "hint":    _SOURCE_RESEARCH_HINT,
            "output_format": _SOURCE_RESEARCH_OUTPUT,
        },
        "scrape": {
            "label": "Scrape-Agent",
            "persona": _SEARCH_PERSONA_BASE + " Du spezialisierst dich auf öffentlich zugängliche Webseiten von Behörden, Verbänden und Institutionen.",
            "method": "Du rufst jede URL direkt ab (HTTP GET), analysierst den HTML-Inhalt und extrahierst relevante neue Dokumente, Meldungen oder Änderungen. Vergleiche mit bekannten Inhalten und melde nur tatsächlich neue Informationen. Wenn die gegebene URL eine Homepage ist (kein direkter Endpunkt wie /news, /aktuelles, /beschluesse), suche zunächst selbstständig nach den relevanten Unterseiten (Newsbereich, Meldungen, Beschlüsse, Pressemitteilungen) und crawle diese — nicht die Homepage selbst.",
            "hint": "Achte besonders auf Neuigkeiten, Pressemitteilungen und aktuelle Meldungen. Deprioritiere redaktionelle Artikel ohne konkreten regulatorischen Gehalt. Melde nur Inhalte, die seit der letzten Recherche neu erschienen sind. **Wichtig:** Deine Aufgabe ist sammeln, nicht bewerten. Nimm im Zweifel auf — Relevanzentscheidungen trifft der nachgelagerte Bewertungs-Agent. Nur offensichtlich themenfremde Inhalte (Sportnachrichten, Stellenanzeigen ohne regulatorischen Bezug) weglassen."
        },
        "fletcher_feed": {
            "label": "Feed-Fetcher",
            "beschreibung": (
                "Fletcher ruft deterministisch alle RSS- und Atom-Feeds aus dem Quellenregister ab. "
                "Er erkennt neue Einträge anhand gespeicherter Entry-IDs (State-File), speichert Rohdaten "
                "(Titel, URL, Datum, Originaltext) und schreibt ein vollständiges Crawl-Receipt — "
                "welche Feeds OK waren, welche Fehler zurückgaben und wie viele neue Einträge gefunden wurden. "
                "Fletcher bewertet, filtert und formuliert nichts — das ist Lars' Aufgabe."
            ),
            "konfiguration": (
                "User-Agent: Branchenradar-Fletcher/1.0\n"
                "State-File: data/fetcher_feed_state.json (bis zu 500 bekannte IDs pro Feed-URL)\n"
                "Output: data/fetcher_feed_latest.json\n"
                "Textvorschau: max. 600 Zeichen (HTML-Tags werden entfernt)\n"
                "Fehlerbehandlung: Quelle wird im Crawl-Receipt mit Fehlertext protokolliert, Lauf läuft weiter"
            ),
        },
        "frank_scrape": {
            "label": "Scrape-Fetcher",
            "beschreibung": (
                "Frank ruft deterministisch alle Scrape-Endpunkte aus dem Quellenregister ab. "
                "Er nutzt die von Maja und Nils hinterlegten CSS-Selektoren, um Artikel-Links auf "
                "Übersichtsseiten zu extrahieren. Neue Links werden anhand eines State-Files erkannt "
                "(bekannte URLs pro Endpunkt) und mit Titel und Snippet an Finn weitergegeben. "
                "Kein KI-Ermessen beim Sammeln — vollständige Abdeckung, auditierbar pro Quelle."
            ),
            "konfiguration": (
                "User-Agent: Branchenradar-Frank/1.0\n"
                "State-File: data/fetcher_scrape_state.json (bis zu 500 bekannte URLs pro Endpunkt)\n"
                "Output: data/fetcher_scrape_latest.json\n"
                "Fallback ohne Selektor: alle internen <a>-Tags mit pfadähnlicher URL\n"
                "Fehlerbehandlung: Quelle wird im Crawl-Receipt mit Fehlertext protokolliert, Lauf läuft weiter"
            ),
        },
        "sol_selector": {
            "label": "Selektor-Finder",
            "beschreibung": (
                "Sol besucht automatisch alle Scrape-Endpunkte im Quellenregister, die noch keinen "
                "CSS-Selektor haben. Er ruft die Seite serverseitig ab, analysiert die HTML-Linkstruktur "
                "und schreibt den besten Kandidaten direkt in sources.json. "
                "Kein manueller Aufwand, kein Raten — Sol erkennt Artikel-Link-Muster deterministisch."
            ),
            "timeout":      "8",
            "min_links":    "3",
            "max_per_run":  "30",
        },
        "feed": {
            "label": "Feed-Agent (Lars)",
            "persona": _SEARCH_PERSONA_BASE + " Du spezialisierst dich auf die inhaltliche Auswertung von Feed-Einträgen, die Fletcher (der Feed-Fetcher) für dich gesammelt hat.",
            "method": (
                "Du erhältst den Fletcher-Export als Input: eine strukturierte Liste neuer Feed-Einträge "
                "mit Titel, Quelle, URL, Datum und Originaltext. Deine Aufgabe ist ausschließlich die "
                "inhaltliche Bewertung dieser Einträge — du rufst keine URLs selbst ab. "
                "Prüfe jeden Eintrag auf Relevanz für Gesundheitseinrichtungen in Deutschland "
                "und erstelle für relevante Einträge Signal-Cards im vorgegebenen Ausgabeformat."
            ),
            "hint": (
                "Fokussiere auf Themen wie Abrechnung, Datenschutz, Hygiene, Qualitätsmanagement "
                "und regulatorische Änderungen. Deprioritiere reine Veranstaltungshinweise ohne "
                "regulatorischen Inhalt. **Wichtig:** Deine Aufgabe ist bewerten und formulieren — "
                "das Sammeln hat Fletcher bereits übernommen. Ein Eintrag pro Signal, "
                "auch wenn mehrere Quellen dasselbe Ereignis berichten."
            ),
        },
        "pdf": {
            "label": "PDF-Agent",
            "persona": _SEARCH_PERSONA_BASE + " Du spezialisierst dich auf verlinkte Dokumente (PDFs, Word-Dateien) von Quellen-Webseiten.",
            "method": "Du durchsuchst die Quellen-URLs nach verlinkten Dokumenten, lädst diese herunter und extrahierst relevante Inhalte aus PDFs und anderen Dokumentformaten. Wenn die gegebene URL eine Übersichtsseite ist, navigiere zunächst zu den spezifischen Dokumentverzeichnissen oder Beschlusslisten — crawle nicht die Homepage selbst.",
            "hint": "Priorisiere aktuelle Leitlinien, Rundschreiben, Beschlüsse und offizielle Bekanntmachungen. Achte auf Versionsnummern und Datumsangaben, um neue von bekannten Dokumenten zu unterscheiden. Ignoriere unveränderte Dokumente. **Wichtig:** Deine Aufgabe ist sammeln, nicht bewerten. Nimm im Zweifel auf — Relevanzentscheidungen trifft der nachgelagerte Bewertungs-Agent. Nur offensichtlich themenfremde Dokumente weglassen."
        },
        "search": {
            "label": "Search-Agent",
            "persona": _SEARCH_PERSONA_BASE + " Du spezialisierst dich auf Quellen mit Login-Pflicht, Datenbanken oder kostenpflichtigen Inhalten.",
            "method": "Du verwendest gespeicherte Zugangsdaten oder öffentliche Suchfunktionen, um in geschützten Bereichen oder Datenbanken nach neuen relevanten Inhalten zu suchen. Wenn die gegebene URL eine Einstiegsseite ist, navigiere selbstständig zu den relevanten Inhaltsbereichen (Beschlüsse, Rundschreiben, Neuigkeiten).",
            "hint": "Falls kein direkter Zugang möglich ist, suche nach öffentlichen Zusammenfassungen, Pressemitteilungen oder alternativen Zugangswegen zur gleichen Information. **Wichtig:** Deine Aufgabe ist sammeln, nicht bewerten. Nimm im Zweifel auf — Relevanzentscheidungen trifft der nachgelagerte Bewertungs-Agent. Nur offensichtlich themenfremde Inhalte weglassen."
        },
        "group": {
            "label": "Gruppierungs-Agent",
            "date_from": "", "date_to": "",
            "last_date_from": "", "last_date_to": "",
            "persona": (
                "Du bist ein redaktioneller Gruppierungs-Agent für den Branchenradar Gesundheitswesen. "
                "Du erhältst zwei Input-Datenbanken als Excel-Dateien: "
                "(1) neue Signale, die von den Such-Agenten in diesem Recherche-Zyklus gefunden wurden, und "
                "(2) alle Signale aus dem letzten Branchenradar-Bericht. "
                "Deine Aufgabe ist es, diese Datenbanken zu konsolidieren und ein aktualisiertes, bereinigtes Signal-Set zu erstellen."
            ),
            "method": (
                "Schritt 1 – Neue Signale auf Duplikate prüfen:\n"
                "Prüfe alle neuen Signale aus den Such-Agenten auf inhaltliche Überschneidungen. "
                "Signale, die exakt dasselbe regulatorische Ereignis beschreiben (gleiche Maßnahme, gleiche Empfänger, gleiches Inkrafttreten), "
                "werden zu einem Signal zusammengefasst. "
                "Das zusammengeführte Signal erhält alle Quellen als Quelle 1 / Quelle 2 / Quelle 3 etc. "
                "Behalte das früheste Veröffentlichungsdatum als primäres Datum.\n\n"
                "**Wichtig: Nicht zusammenfassen**, wenn sich Empfänger, Handlung, Frist oder Inkrafttreten unterscheiden — "
                "auch wenn Signale aus derselben Quelle oder demselben Gesetz stammen. "
                "Beispiel: Spargesetz mit Arzneimittelausschluss (sofort, alle GKV-Praxen) und Wegfall Hygienezuschlag "
                "(ab 2027, nur bestimmte Einrichtungen) = zwei separate Signale, nicht eines.\n\n"
                "Schritt 2 – Abgleich mit bestehendem Reporting:\n"
                "Vergleiche die konsolidierten neuen Signale mit den Signalen aus dem letzten Branchenradar-Bericht. "
                "Entscheide für jedes neue Signal:\n"
                "– Ist es ein Update, eine Ergänzung oder Konkretisierung eines bestehenden Signals? "
                "→ Status: UPDATE. Hänge die neuen Informationen als separaten Absatz an die bestehende Zusammenfassung an, "
                "versehen mit dem Datum der Neuerung im Format [TT.MM.JJJJ]: …\n"
                "– Ist es ein völlig neues Thema, das bisher nicht im Branchenradar war? "
                "→ Status: NEU.\n"
                "– Bestehende Signale aus dem alten Bericht ohne jede Neuerung erhalten Status: leer (unverändert). "
                "Sie werden unverändert in die Ausgabe übernommen."
            ),
            "hint": (
                "Nutze semantisches Verständnis für den Abgleich — nicht nur Stichwortübereinstimmungen. "
                "Verschiedene Meldungen zur selben Gesetzgebung, zum selben Beschluss oder zum selben Verfahren gehören zusammen — "
                "aber nur wenn sie dieselbe Maßnahme mit denselben Empfängern und demselben Inkrafttreten beschreiben.\n\n"
                "Faustregel für Zusammenfassen vs. Trennen:\n"
                "– Gleiche Maßnahme, gleiche Empfänger, gleiches Inkrafttreten, mehrere Quellen → zusammenfassen, alle Quellen nennen\n"
                "– Gleicher Artikel/Beschluss, aber unterschiedliche Regelungen darin → getrennte Signale behalten\n"
                "– Regionale Umsetzung einer bundesweiten Regelung (z.B. KV Nordrhein setzt um was G-BA beschlossen hat) "
                "→ zusammenfassen, Primärquelle = G-BA, regionale Quelle als Quelle 2\n\n"
                "Achte besonders auf Datumsangaben: neuere Meldungen zum gleichen Thema sind Update-Kandidaten. "
                "Im Zweifel lieber zwei Signale behalten als eines zu verwerfen. "
                "Behalte Originalzusammenfassungen bestehender Signale vollständig — ergänze nur, lösche nicht."
            ),
            "output_format": (
                "## Ausgabeformat – Signale (MATCH)\n\n"
                "Erstelle eine tabellarische Ausgabe im folgenden Format. Jede Zeile beschreibt eine Gruppe zusammengehöriger Signale.\n\n"
                "**Spalten der MATCH-Tabelle:**\n"
                "- Spalte 1 (sig1): ID des primären Signals aus Signale (RAW) — das Signal, das die Gruppe anführt\n"
                "- Spalte 2 (sig2): ID eines weiteren Signals, das mit sig1 zusammengeführt wird (leer = keine weiteren)\n"
                "- Spalte 3 (sig2_type): NEW wenn sig2 ein neu recherchiertes Signal ist | OLD wenn sig2 aus dem letzten Reporting stammt\n"
                "- Spalte 4 (sig3): ID eines dritten Signals (optional)\n"
                "- Spalte 5 (sig3_type): NEW | OLD\n"
                "- Spalte 6 (sig4): ID eines vierten Signals (optional)\n"
                "- Spalte 7 (sig4_type): NEW | OLD\n"
                "- Spalte 8 (sig5): ID eines fünften Signals (optional)\n"
                "- Spalte 9 (sig5_type): NEW | OLD\n\n"
                "**Signal-IDs** haben das Format JJMMDDXXXXX (z.B. 260806000042). "
                "Jedes Signal aus Signale (RAW) hat eine solche ID. Bestehende Signale aus dem letzten Reporting haben ebenfalls IDs in diesem Format.\n\n"
                "**Beispielzeile:**\n"
                "sig1=260806000017 | sig2=260806000031 | sig2_type=NEW | sig3=260801000008 | sig3_type=OLD\n\n"
                "Signale ohne Überschneidung mit anderen Signalen erscheinen als eigenständige Zeilen mit nur sig1 befüllt (alle anderen Spalten leer).\n\n"
                "Nach der MATCH-Tabelle: Erstelle zusätzlich eine vollständige Signal-Liste für die redaktionelle Weiterbearbeitung:\n\n"
                "**Status:** NEU | UPDATE | (leer = unverändert)\n"
                "**Titel:** [Sprechender Action-Titel]\n"
                "**Zusammenfassung:** [Bestehende Zusammenfassung. Bei UPDATE: Neuer Absatz mit [TT.MM.JJJJ]: Neue Informationen]\n"
                "**Datum:** [Datum des Signals bzw. aktuellstes Datum bei Updates]\n"
                "**Entwicklungsstand:** [Beobachtung | Veröffentlicht | Beschlossen | In Kraft]\n"
                "**Quelle 1:** [Name] | **Quellenlink 1:** [URL]\n"
                "**Quelle 2:** [Name] | **Quellenlink 2:** [URL] (falls vorhanden)\n"
                "**Quelle 3:** [Name] | **Quellenlink 3:** [URL] (falls vorhanden)\n\n"
                "---"
            )
        },
        "bewertung": {
            "label": "Bewertungs-Agent",
            "persona": (
                "Du bist ein redaktioneller Bewertungs-Agent für den Branchenradar Gesundheitswesen. "
                "Du erhältst drei Input-Datenbanken: "
                "(1) Signale (RAW) – neue, von den Such-Agenten recherchierte Rohdaten, "
                "(2) Signale (FINAL) – die konsolidierten Signale des letzten Branchenradar-Reportings, "
                "(3) Signale (MATCH) – die Zuordnungstabelle des Gruppierungs-Agenten, die festlegt, "
                "welche RAW-Signale zusammengehören und ob sie NEU oder ein UPDATE zu einem bestehenden FINAL-Signal sind. "
                "Deine Aufgabe ist es, auf dieser Basis den vollständig angereicherten, neuen Signaldatensatz zu erstellen, "
                "der als nächstes Signale (FINAL) gespeichert wird."
            ),
            "method": (
                "Schritt 1 – Signale zusammenführen:\n"
                "Nutze die MATCH-Tabelle als Fahrplan. Für jede MATCH-Gruppe:\n"
                "– Kombiniere alle referenzierten RAW-Signale (NEW) zu einem inhaltlich kohärenten Signal. "
                "Führe Titel, Zusammenfassung und Quellenangaben zusammen.\n"
                "– Falls die Gruppe auch OLD-Signale enthält (aus FINAL), erweitere das bestehende FINAL-Signal "
                "um die neuen Informationen. Hänge neue Absätze mit Datumsmarkierung [TT.MM.JJJJ] an.\n"
                "– Bestehende FINAL-Signale ohne jede MATCH-Verbindung werden unverändert übernommen.\n\n"
                "Schritt 2 – Eigenständige neue Signale:\n"
                "RAW-Signale, die in keiner MATCH-Gruppe auftauchen, sind völlig neue Themen. "
                "Übernimm sie direkt als neue Signale in den Output.\n\n"
                "Schritt 3 – Anreicherung aller Signale:\n"
                "Für jedes Signal – ob neu, aktualisiert oder unverändert – vergib oder überprüfe:\n"
                "– **Priorität:** MUSS (unmittelbarer Handlungsbedarf, rechtlich verpflichtend oder hohe finanzielle Auswirkung), "
                "SOLLTE (relevante Entwicklung, Handlung empfohlen), "
                "KANN (interessant zur Beobachtung, noch kein konkreter Handlungsbedarf)\n"
                "– **Handlungsempfehlung:** Ein konkreter, umsetzbarer Handlungsschritt im Imperativ (z.B. 'TI-Anbindung bis Q2 prüfen und beauftragen'). "
                "Nicht allgemein, sondern spezifisch für Gesundheitseinrichtungen.\n"
                "– **Umsetzungszeit:** Realistische Einschätzung des Zeitaufwands für eine typische Praxis "
                "(z.B. '1–2 Stunden', '1 Tag', '1–2 Wochen', '1–3 Monate').\n"
                "– **Betroffene Rollen:** Wer in der Praxis ist konkret betroffen? "
                "Auswahl aus: Praxisleitung, Ärzt:innen, MFA / ZFA, Abrechnung, Praxismanagement, Datenschutz, QM\n"
                "– **Handlungszeitpunkt:** Sofort | Kurzfristig (< 3 Monate) | Mittelfristig (3–12 Monate) | Langfristig (> 12 Monate) | Beobachten\n"
                "– **Kategorie:** Krankenkassen & GKV | Digitalisierung & TI | Gesetze & Regulatorien | Personal & Tarife | Praxismanagement\n"
                "– **Reporting-Status:** NEU | UPDATE | (leer = unverändert)"
            ),
            "hint": (
                "Sei bei der Prioritätsvergabe streng: MUSS ist nur für wirklich zwingende, "
                "fristgebundene oder sanktionsbewehrte Verpflichtungen. Nicht jede neue Entwicklung ist ein MUSS. "
                "Die Handlungsempfehlung soll dem Praxisinhaber oder -manager sofort klar machen, was jetzt konkret zu tun ist – "
                "keine vagen Formulierungen wie 'informieren' oder 'beobachten' bei MUSS-Signalen. "
                "Betroffene Rollen sparsam vergeben – nur die Rollen, die tatsächlich direkt handeln müssen."
            ),
            "output_format": (
                "## Ausgabeformat – Signale (FINAL)\n\n"
                "Erstelle eine vollständige Liste aller Signale im folgenden Format. "
                "Jedes Signal auf einer neuen Sektion, getrennt durch ---\n\n"
                "**ID:** [Übernehmen aus RAW/FINAL – bei vollständig neuen Signalen ohne MATCH: neue ID aus RAW]\n"
                "**Reporting-Status:** NEU | UPDATE | (leer)\n"
                "**Priorität:** MUSS | SOLLTE | KANN\n"
                "**Titel:** [Sprechender Action-Titel]\n"
                "**Zusammenfassung:** [Vollständige Zusammenfassung. Bei UPDATE: bestehende Zusammenfassung + [TT.MM.JJJJ]: neue Informationen]\n"
                "**Kategorie:** [Krankenkassen & GKV | Digitalisierung & TI | Gesetze & Regulatorien | Personal & Tarife | Praxismanagement]\n"
                "**Entwicklungsstand:** [Beobachtung | Veröffentlicht | Beschlossen | In Kraft]\n"
                "**Handlungsempfehlung:** [Konkreter Handlungsschritt im Imperativ]\n"
                "**Umsetzungszeit:** [z.B. 1–2 Stunden | 1 Tag | 1–2 Wochen | 1–3 Monate]\n"
                "**Betroffene Rollen:** [Kommagetrennte Liste aus: Praxisleitung, Ärzt:innen, MFA / ZFA, Abrechnung, Praxismanagement, Datenschutz, QM]\n"
                "**Handlungszeitpunkt:** [Sofort | Kurzfristig (< 3 Monate) | Mittelfristig (3–12 Monate) | Langfristig (> 12 Monate) | Beobachten]\n"
                "**Datum:** [Datum des Signals / aktuellstes Datum bei Updates]\n"
                "**Quelle 1:** [Name] | **Quellenlink 1:** [URL]\n"
                "**Quelle 2:** [Name] | **Quellenlink 2:** [URL] (falls vorhanden)\n"
                "**Quelle 3:** [Name] | **Quellenlink 3:** [URL] (falls vorhanden)\n\n"
                "---"
            )
        }
    }
}

def load_agent_configs():
    import copy
    cfg = copy.deepcopy(DEFAULT_AGENT_CONFIGS)
    if not os.path.exists(AGENT_CONFIGS_FILE):
        return cfg
    data = _load(AGENT_CONFIGS_FILE)
    # Only override with saved value if it's non-empty
    if data.get("output_format", "").strip():
        cfg["output_format"] = data["output_format"]
    for key in cfg["agents"]:
        saved = data.get("agents", {}).get(key, {})
        for field in ("persona", "method", "hint", "output_format", "beschreibung", "konfiguration"):
            if saved.get(field, "").strip():
                cfg["agents"][key][field] = saved[field]
        for field in ("date_from", "date_to", "last_date_from", "last_date_to",
                      "timeout", "min_links", "max_per_run",
                      "schedule_type", "schedule_day", "schedule_weekday",
                      "schedule_time", "schedule_enabled", "maja_model"):
            if field in saved:
                cfg["agents"][key][field] = saved[field]
    return cfg

def save_agent_configs(d): _save(AGENT_CONFIGS_FILE, d)

def load_settings():
    if not os.path.exists(SETTINGS_FILE):
        return {"default_todo_categories": DEFAULT_TODO_CATEGORIES}
    with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_settings(d): _save(SETTINGS_FILE, d)

def load_agent_reports():
    if not os.path.exists(AGENT_REPORTS_FILE):
        return []
    return _load(AGENT_REPORTS_FILE) or []

def save_agent_reports(data): _save(AGENT_REPORTS_FILE, data)

def load_signal_matches():
    if not os.path.exists(SIGNAL_MATCHES_FILE):
        return []
    return _load(SIGNAL_MATCHES_FILE) or []

def save_signal_matches(data): _save(SIGNAL_MATCHES_FILE, data)

def load_signal_final():
    if not os.path.exists(SIGNAL_FINAL_FILE):
        return {"timestamp": None, "signals": []}
    data = _load(SIGNAL_FINAL_FILE)
    if isinstance(data, list):
        return {"timestamp": None, "signals": data}
    return data

def save_signal_final(signals, timestamp):
    _save(SIGNAL_FINAL_FILE, {"timestamp": timestamp, "signals": signals})

def archive_signal_final():
    if not os.path.exists(SIGNAL_FINAL_FILE):
        return
    os.makedirs(SIGNAL_FINAL_ARCHIVE_DIR, exist_ok=True)
    data = load_signal_final()
    ts = data.get("timestamp") or datetime.now().strftime("%y%m%d%H%M")
    archive_path = os.path.join(SIGNAL_FINAL_ARCHIVE_DIR, f"{ts}.json")
    _save(archive_path, data)

def list_signal_final_archives():
    if not os.path.exists(SIGNAL_FINAL_ARCHIVE_DIR):
        return []
    files = sorted(
        [f[:-5] for f in os.listdir(SIGNAL_FINAL_ARCHIVE_DIR) if f.endswith(".json")],
        reverse=True
    )
    return files

def load_signal_final_archive(ts):
    path = os.path.join(SIGNAL_FINAL_ARCHIVE_DIR, f"{ts}.json")
    if not os.path.exists(path):
        return None
    return _load(path)

def next_signal_raw_id():
    cfg = load_settings()
    counter = cfg.get("signal_raw_counter", 0) + 1
    cfg["signal_raw_counter"] = counter
    save_settings(cfg)
    return f"{datetime.now().strftime('%y%m%d')}{counter:05d}"

def entity_todo_categories(entity_id):
    if not entity_id:
        return DEFAULT_TODO_CATEGORIES
    e = get_entity(entity_id)
    if e:
        return e.get("todo_categories", DEFAULT_TODO_CATEGORIES)
    return DEFAULT_TODO_CATEGORIES

def sections_by_id():
    return {s["id"]: s for s in load_sections()}

def get_entity(eid):
    return next((e for e in load_entities() if e["id"] == eid), None)

def get_user(uid):
    return next((u for u in load_users() if u["id"] == uid), None)

# ── CSRF ─────────────────────────────────────────────────────────────────────

def _csrf_token():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)
    return session["csrf_token"]

def _csrf_check():
    token = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
    if not token or not secrets.compare_digest(token, session.get("csrf_token", "")):
        abort(403)

app.jinja_env.globals["csrf_token"] = _csrf_token

_CSRF_EXEMPT_PREFIXES = ("/admin/maja/", "/admin/selector/run", "/agents/save",
                         "/deploy-webhook")

@app.before_request
def csrf_protect():
    if request.method not in ("POST", "PUT", "DELETE", "PATCH"):
        return
    if any(request.path.startswith(p) for p in _CSRF_EXEMPT_PREFIXES):
        return
    if request.is_json:
        return
    _csrf_check()

# ── Login-Drosselung ──────────────────────────────────────────────────────────

_login_attempts: dict = {}  # ip -> [timestamp, ...]
_MAX_ATTEMPTS = 5
_LOCKOUT_S    = 60

def _check_rate_limit(ip):
    now = datetime.now().timestamp()
    attempts = [t for t in _login_attempts.get(ip, []) if now - t < _LOCKOUT_S]
    _login_attempts[ip] = attempts
    if len(attempts) >= _MAX_ATTEMPTS:
        return False
    return True

def _record_attempt(ip):
    _login_attempts.setdefault(ip, []).append(datetime.now().timestamp())

def _clear_attempts(ip):
    _login_attempts.pop(ip, None)

def _safe_next(next_url):
    if next_url and next_url.startswith("/") and not next_url.startswith("//"):
        return next_url
    return None

def _hash(pw):
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt(rounds=12)).decode()

def _verify(pw, stored_hash):
    try:
        return bcrypt.checkpw(pw.encode(), stored_hash.encode())
    except Exception:
        return False


# ── Auth decorators ───────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not session.get("user_id"):
                return redirect(url_for("login", next=request.path))
            if session.get("role") not in roles:
                abort(403)
            return f(*args, **kwargs)
        return decorated
    return decorator


# ── Context processor ────────────────────────────────────────────────────────

@app.context_processor
def inject_globals():
    uid  = session.get("user_id")
    eid  = session.get("entity_id")
    role = session.get("role")

    entity_users = []
    if role in ("admin", "superadmin") and eid:
        entity_users = [u for u in load_users() if u.get("entity_id") == eid]

    badge = 0
    if uid:
        todos = load_todos()
        if role in ("admin", "superadmin"):
            # open todos for entity + unread done-notifications
            badge = sum(
                1 for t in todos
                if t.get("entity_id") == eid and (
                    (not t.get("done")) or
                    (t.get("done") and t.get("assigned_by") == uid and not t.get("admin_read_at"))
                )
            )
        else:
            # user: unread assigned todos
            badge = sum(1 for t in todos if t.get("assigned_to") == uid and not t.get("assignee_read_at"))

    return {
        "todo_categories": entity_todo_categories(eid),
        "pending_todos":   badge,
        "entity_users":    entity_users,
    }


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        ip       = request.remote_addr
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if not _check_rate_limit(ip):
            flash("Zu viele Fehlversuche. Bitte 60 Sekunden warten.", "error")
            return render_template("login.html")
        users = load_users()
        user  = next((u for u in users if u["username"] == username), None)
        if user and _verify(password, user["pass_hash"]):
            _clear_attempts(ip)
            session.clear()
            session["user_id"]   = user["id"]
            session["username"]  = user["username"]
            session["role"]      = user["role"]
            session["name"]      = user.get("name", username)
            session["entity_id"] = user.get("entity_id")
            next_url = _safe_next(request.args.get("next"))
            return redirect(next_url or url_for("dashboard"))
        _record_attempt(ip)
        flash("Benutzername oder Passwort falsch.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("dashboard"))


# ── Public dashboard ──────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    all_signals = load_signals()
    sections    = load_sections()
    sec_map     = {s["id"]: s for s in sections}

    # Determine which sections the current user's entity has access to
    entity_id   = session.get("entity_id")
    role        = session.get("role")
    if role == "superadmin" or not entity_id:
        allowed_section_ids = None  # sees everything
    else:
        entity = get_entity(entity_id)
        allowed_section_ids = set(entity.get("section_ids", [])) if entity else set()

    # Filter signals to allowed sections
    if allowed_section_ids is not None:
        signals = [s for s in all_signals
                   if any(sid in allowed_section_ids for sid in s.get("section_ids", []))]
        visible_sections = [s for s in sections if s["id"] in allowed_section_ids]
    else:
        signals = all_signals
        visible_sections = sections

    # For "user" role: additionally filter by team categories
    if role == "user" and entity_id:
        entity = get_entity(entity_id)
        if entity:
            uid = session.get("user_id")
            user_teams = [t for t in entity.get("teams", []) if uid in t.get("member_ids", [])]
            if user_teams:
                team_cats = set()
                for tm in user_teams:
                    team_cats.update(tm.get("category_ids", []))
                if team_cats:
                    signals = [s for s in signals if s.get("category") in team_cats]

    prio_order = ["muss", "sollte", "kann"]
    def sort_key(s):
        try:   pi = prio_order.index(s.get("priority", "sollte"))
        except ValueError: pi = 1
        return (pi, s.get("date", ""))
    signals.sort(key=sort_key)

    prio_counts = {p: 0 for p in prio_order}
    for sig in signals:
        p = sig.get("priority", "sollte")
        if p in prio_counts:
            prio_counts[p] += 1

    sec_counts = {}
    for sig in signals:
        for sid in sig.get("section_ids", []):
            sec_counts[sid] = sec_counts.get(sid, 0) + 1

    return render_template(
        "dashboard.html",
        signals=signals, prio_counts=prio_counts,
        categories=CATEGORIES, status_labels=STATUS_LABELS,
        priority_labels=SIGNAL_PRIORITY_LABELS,
        entwicklungsstand_labels=ENTWICKLUNGSSTAND_LABELS,
        handlungszeitpunkt_labels=HANDLUNGSZEITPUNKT_LABELS,
        aufwand_labels=AUFWAND_LABELS,
        sections=visible_sections, sec_map=sec_map, sec_counts=sec_counts,
        now=datetime.now().strftime("%d. %B %Y, %H:%M"),
        total=len(signals),
    )


# ── Superadmin: Entity management ─────────────────────────────────────────────

@app.route("/superadmin")
@role_required("superadmin")
def sa_dashboard():
    entities  = load_entities()
    users     = load_users()
    sections  = load_sections()
    sec_map   = {s["id"]: s for s in sections}
    counts = {}
    for u in users:
        eid = u.get("entity_id")
        if eid:
            counts[eid] = counts.get(eid, 0) + 1
    return render_template("sa_dashboard.html",
        entities=entities, user_counts=counts, users=users,
        sections=sections, sec_map=sec_map, role_labels=ROLE_LABELS)

@app.route("/superadmin/entities/<eid>/edit", methods=["GET", "POST"])
@role_required("superadmin")
def sa_entity_edit(eid):
    entities = load_entities()
    entity = next((e for e in entities if e["id"] == eid), None)
    if not entity:
        abort(404)
    if request.method == "POST":
        entity["name"]        = request.form.get("name", "").strip()
        entity["section_ids"] = request.form.getlist("section_ids")
        save_entities(entities)
        flash("Entity aktualisiert.", "success")
        return redirect(url_for("sa_dashboard"))
    return render_template("sa_entity_form.html", entity=entity,
        sections=load_sections())

@app.route("/superadmin/entities/new", methods=["GET", "POST"])
@role_required("superadmin")
def sa_entity_new():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Name ist erforderlich.", "error")
            return render_template("sa_entity_form.html", entity=None, sections=load_sections())
        entities = load_entities()
        cfg = load_settings()
        default_cats = cfg.get("default_todo_categories", DEFAULT_TODO_CATEGORIES)
        new_entity = {
            "id":              str(uuid.uuid4()),
            "name":            name,
            "created_at":      datetime.now().strftime("%Y-%m-%d"),
            "section_ids":     request.form.getlist("section_ids"),
            "todo_categories": default_cats,
            "teams": [
                {"id": str(uuid.uuid4()), "name": tn, "member_ids": [], "category_ids": []}
                for tn in cfg.get("default_team_names", ["Team Abrechnung", "Team Personal"])
            ],
        }
        entities.append(new_entity)
        save_entities(entities)

        # Optionally create admin user right away
        admin_username = request.form.get("admin_username", "").strip()
        admin_password = request.form.get("admin_password", "").strip()
        admin_name     = request.form.get("admin_name", "").strip()
        if admin_username and admin_password:
            users = load_users()
            if any(u["username"] == admin_username for u in users):
                flash(f"Entity '{name}' angelegt. Benutzername '{admin_username}' existiert bereits — Admin manuell anlegen.", "error")
            else:
                users.append({
                    "id":        str(uuid.uuid4()),
                    "username":  admin_username,
                    "pass_hash": _hash(admin_password),
                    "role":      "admin",
                    "entity_id": new_entity["id"],
                    "name":      admin_name or admin_username,
                })
                save_users(users)
                flash(f"Entity '{name}' und Admin '{admin_username}' angelegt.", "success")
        else:
            flash(f"Entity '{name}' angelegt. Admin noch hinzufügen.", "success")
        return redirect(url_for("sa_dashboard"))
    return render_template("sa_entity_form.html", entity=None, sections=load_sections())

@app.route("/superadmin/entities/<eid>/delete", methods=["POST"])
@role_required("superadmin")
def sa_entity_delete(eid):
    entities = [e for e in load_entities() if e["id"] != eid]
    save_entities(entities)
    # Remove users of this entity
    users = [u for u in load_users() if u.get("entity_id") != eid]
    save_users(users)
    flash("Entity und zugehörige Nutzer gelöscht.", "success")
    return redirect(url_for("sa_dashboard"))

@app.route("/superadmin/users/new", methods=["GET", "POST"])
@role_required("superadmin")
def sa_user_new():
    if request.method == "POST":
        return _create_user(request.form, redirect_to=url_for("sa_dashboard"))
    entities = load_entities()
    return render_template("user_form.html", user=None, entities=entities,
        role_labels=ROLE_LABELS, action=url_for("sa_user_new"),
        show_all_roles=True)

@app.route("/superadmin/users/<uid>/delete", methods=["POST"])
@role_required("superadmin")
def sa_user_delete(uid):
    users = [u for u in load_users() if u["id"] != uid]
    save_users(users)
    flash("Nutzer gelöscht.", "success")
    return redirect(url_for("sa_dashboard"))

@app.route("/superadmin/users/<uid>/edit", methods=["GET", "POST"])
@role_required("superadmin")
def sa_user_edit(uid):
    users = load_users()
    user = next((u for u in users if u["id"] == uid), None)
    if not user:
        abort(404)
    if request.method == "POST":
        user["name"]      = request.form.get("name", "").strip()
        user["username"]  = request.form.get("username", "").strip()
        user["role"]      = request.form.get("role", "user")
        user["entity_id"] = request.form.get("entity_id") or None
        pw = request.form.get("password", "").strip()
        if pw:
            user["pass_hash"] = _hash(pw)
        save_users(users)
        flash("Nutzer aktualisiert.", "success")
        return redirect(url_for("sa_dashboard"))
    entities = load_entities()
    return render_template("user_form.html", user=user, entities=entities,
        role_labels=ROLE_LABELS, action=url_for("sa_user_edit", uid=uid),
        show_all_roles=True)


# ── Admin: User management within entity ──────────────────────────────────────

@app.route("/admin")
@role_required("admin", "superadmin")
def admin_dashboard():
    # Superadmin → redirect to sa_dashboard
    if session["role"] == "superadmin":
        return redirect(url_for("sa_dashboard"))
    entity_id = session["entity_id"]
    entity = get_entity(entity_id)
    users = [u for u in load_users() if u.get("entity_id") == entity_id]
    return render_template("admin_users.html",
        entity=entity, users=users, role_labels=ROLE_LABELS)

@app.route("/admin/users/new", methods=["GET", "POST"])
@role_required("admin", "superadmin")
def admin_user_new():
    entity_id = session["entity_id"]
    if request.method == "POST":
        form = request.form.copy()
        # Admin can only create users in their own entity
        return _create_user(form, redirect_to=url_for("admin_dashboard"),
                            force_entity=entity_id, force_role="user")
    entity = get_entity(entity_id)
    return render_template("user_form.html", user=None,
        entities=[entity] if entity else [],
        role_labels={"user": "Nutzer"},
        action=url_for("admin_user_new"), show_all_roles=False)

@app.route("/admin/users/<uid>/delete", methods=["POST"])
@role_required("admin", "superadmin")
def admin_user_delete(uid):
    entity_id = session["entity_id"]
    users = load_users()
    target = next((u for u in users if u["id"] == uid), None)
    if not target or target.get("entity_id") != entity_id:
        abort(403)
    users = [u for u in users if u["id"] != uid]
    save_users(users)
    flash("Nutzer entfernt.", "success")
    return redirect(url_for("admin_dashboard"))


# ── Signal management (admin + superadmin) ────────────────────────────────────

@app.route("/admin/signals/raw")
@role_required("admin", "superadmin")
def signal_list_raw():
    return signal_list()

@app.route("/admin/signals")
@role_required("admin", "superadmin")
def signal_list():
    signals = load_signals()
    signals.sort(key=lambda s: s.get("date", ""), reverse=True)
    sections = load_sections()
    sec_map  = {s["id"]: s for s in sections}
    return render_template("admin_signals.html",
        signals=signals, categories=CATEGORIES, status_labels=STATUS_LABELS,
        priority_labels=SIGNAL_PRIORITY_LABELS,
        entwicklungsstand_labels=ENTWICKLUNGSSTAND_LABELS,
        handlungszeitpunkt_labels=HANDLUNGSZEITPUNKT_LABELS,
        aufwand_labels=AUFWAND_LABELS,
        signal_rollen=SIGNAL_ROLLEN,
        sections=sections, sec_map=sec_map)

@app.route("/admin/signals/new", methods=["GET", "POST"])
@role_required("admin", "superadmin")
def signal_new():
    if request.method == "POST":
        sig = _signal_from_form(request.form)
        signals = load_signals()
        signals.append(sig)
        save_signals(signals)
        flash("Signal gespeichert.", "success")
        return redirect(url_for("signal_list"))
    return render_template("signal_form.html", signal=None,
        categories=CATEGORIES, status_labels=STATUS_LABELS,
        priority_labels=SIGNAL_PRIORITY_LABELS,
        entwicklungsstand_labels=ENTWICKLUNGSSTAND_LABELS,
        handlungszeitpunkt_labels=HANDLUNGSZEITPUNKT_LABELS,
        aufwand_labels=AUFWAND_LABELS,
        signal_rollen=SIGNAL_ROLLEN, region_options=REGION_OPTIONS,
        sections=load_sections(), action=url_for("signal_new"))

@app.route("/admin/signals/<sig_id>/edit", methods=["GET", "POST"])
@role_required("admin", "superadmin")
def signal_edit(sig_id):
    signals = load_signals()
    sig = next((s for s in signals if s["id"] == sig_id), None)
    if not sig:
        abort(404)
    if request.method == "POST":
        updated = _signal_from_form(request.form, existing_id=sig_id)
        for i, s in enumerate(signals):
            if s["id"] == sig_id:
                signals[i] = updated
                break
        save_signals(signals)
        flash("Signal aktualisiert.", "success")
        return redirect(url_for("signal_list"))
    return render_template("signal_form.html", signal=sig,
        categories=CATEGORIES, status_labels=STATUS_LABELS,
        priority_labels=SIGNAL_PRIORITY_LABELS,
        entwicklungsstand_labels=ENTWICKLUNGSSTAND_LABELS,
        handlungszeitpunkt_labels=HANDLUNGSZEITPUNKT_LABELS,
        aufwand_labels=AUFWAND_LABELS,
        signal_rollen=SIGNAL_ROLLEN, region_options=REGION_OPTIONS,
        sections=load_sections(), action=url_for("signal_edit", sig_id=sig_id))

@app.route("/admin/signals/<sig_id>/delete", methods=["POST"])
@role_required("admin", "superadmin")
def signal_delete(sig_id):
    signals = [s for s in load_signals() if s["id"] != sig_id]
    save_signals(signals)
    flash("Signal gelöscht.", "success")
    return redirect(url_for("signal_list"))

# ── ToDos ─────────────────────────────────────────────────────────────────────

@app.route("/todos")
@login_required
def todos():
    uid  = session.get("user_id")
    eid  = session.get("entity_id")
    role = session.get("role")
    all_todos = load_todos()

    if role in ("admin", "superadmin"):
        my_todos = [t for t in all_todos if t.get("entity_id") == eid]
        # Mark done-notifications as read for this admin
        changed = False
        for t in all_todos:
            if t.get("assigned_by") == uid and t.get("done") and not t.get("admin_read_at"):
                t["admin_read_at"] = datetime.now().isoformat()
                changed = True
        if changed:
            save_todos(all_todos)
    else:
        my_todos = [t for t in all_todos if t.get("assigned_to") == uid]
        # Mark assignee unread as read
        changed = False
        for t in all_todos:
            if t.get("assigned_to") == uid and not t.get("assignee_read_at"):
                t["assignee_read_at"] = datetime.now().isoformat()
                changed = True
        if changed:
            save_todos(all_todos)

    sig_map    = {s["id"]: s for s in load_signals()}
    user_map   = {u["id"]: u for u in load_users()}
    cats       = entity_todo_categories(eid)
    today      = datetime.now().strftime("%Y-%m-%d")
    return render_template("todos.html",
        todos=my_todos, sig_map=sig_map, user_map=user_map,
        categories=cats, today=today, status_labels=STATUS_LABELS)

@app.route("/todos/new", methods=["POST"])
@role_required("admin", "superadmin")
def todo_new():
    uid = session.get("user_id")
    eid = session.get("entity_id")
    assigned_to = request.form.get("assigned_to", "").strip() or None
    todo = {
        "id":              str(uuid.uuid4()),
        "signal_id":       request.form.get("signal_id", ""),
        "signal_title":    request.form.get("signal_title", ""),
        "entity_id":       eid,
        "category":        request.form.get("category", ""),
        "deadline":        request.form.get("deadline", ""),
        "comment":         request.form.get("comment", "").strip(),
        "done":            False,
        "done_at":         None,
        "done_comment":    "",
        "done_by":         None,
        "assigned_to":     assigned_to,
        "assigned_by":     uid if assigned_to else None,
        "assigned_at":     datetime.now().isoformat() if assigned_to else None,
        "assignee_read_at": None,
        "admin_read_at":   None,
        "created_by":      uid,
        "created_at":      datetime.now().isoformat(),
    }
    todos = load_todos()
    todos.append(todo)
    save_todos(todos)
    flash("ToDo angelegt.", "success")
    return redirect(request.form.get("next") or url_for("todos"))

@app.route("/todos/<todo_id>/assign", methods=["POST"])
@role_required("admin", "superadmin")
def todo_assign(todo_id):
    uid = session.get("user_id")
    assigned_to = request.form.get("assigned_to", "").strip() or None
    todos = load_todos()
    for t in todos:
        if t["id"] == todo_id:
            t["assigned_to"]      = assigned_to
            t["assigned_by"]      = uid if assigned_to else None
            t["assigned_at"]      = datetime.now().isoformat() if assigned_to else None
            t["assignee_read_at"] = None
            break
    save_todos(todos)
    return redirect(url_for("todos"))

@app.route("/todos/<todo_id>/complete", methods=["POST"])
@login_required
def todo_complete(todo_id):
    uid = session.get("user_id")
    todos = load_todos()
    for t in todos:
        if t["id"] == todo_id:
            t["done"]          = True
            t["done_at"]       = datetime.now().isoformat()
            t["done_comment"]  = request.form.get("done_comment", "").strip()
            t["done_by"]       = uid
            t["admin_read_at"] = None  # admin gets notified
            break
    save_todos(todos)
    return redirect(url_for("todos"))

@app.route("/todos/<todo_id>/reopen", methods=["POST"])
@role_required("admin", "superadmin")
def todo_reopen(todo_id):
    todos = load_todos()
    for t in todos:
        if t["id"] == todo_id:
            t["done"]         = False
            t["done_at"]      = None
            t["done_comment"] = ""
            t["done_by"]      = None
            t["admin_read_at"] = None
            break
    save_todos(todos)
    return redirect(url_for("todos"))

@app.route("/todos/<todo_id>/delete", methods=["POST"])
@role_required("admin", "superadmin")
def todo_delete(todo_id):
    todos = [t for t in load_todos() if t["id"] != todo_id]
    save_todos(todos)
    return redirect(url_for("todos"))


# ── Settings ──────────────────────────────────────────────────────────────────

@app.route("/settings", methods=["GET", "POST"])
@role_required("admin", "superadmin")
def settings():
    eid = session.get("entity_id")
    entities = load_entities()
    entity = next((e for e in entities if e["id"] == eid), None)
    if not entity:
        abort(404)
    if request.method == "POST":
        action = request.form.get("action", "categories")
        if action == "categories":
            cats = [c.strip() for c in request.form.getlist("categories") if c.strip()]
            entity["todo_categories"] = cats
        elif action == "teams":
            teams = []
            i = 0
            while request.form.get(f"team_{i}_name") is not None:
                name    = request.form.get(f"team_{i}_name", "").strip()
                tid     = request.form.get(f"team_{i}_id") or str(uuid.uuid4())
                members = request.form.getlist(f"team_{i}_members")
                cats_t  = request.form.getlist(f"team_{i}_cats")
                if name:
                    teams.append({"id": tid, "name": name, "member_ids": members, "category_ids": cats_t})
                i += 1
            entity["teams"] = teams
        save_entities(entities)
        flash("Einstellungen gespeichert.", "success")
        return redirect(url_for("settings"))
    cats       = entity.get("todo_categories", DEFAULT_TODO_CATEGORIES)
    teams      = entity.get("teams", [])
    all_users  = [u for u in load_users() if u.get("entity_id") == eid]
    return render_template("settings.html", entity=entity, categories=cats,
        teams=teams, entity_users=all_users, signal_categories=CATEGORIES)

@app.route("/superadmin/functions")
@role_required("superadmin")
def sa_functions():
    return render_template("sa_functions.html",
        now=datetime.now().strftime("%d. %B %Y"))

@app.route("/superadmin/logbook")
@role_required("superadmin")
def sa_logbook():
    return render_template("sa_logbook.html")

@app.route("/superadmin/settings", methods=["GET", "POST"])
@role_required("superadmin")
def sa_settings():
    cfg = load_settings()
    if request.method == "POST":
        section = request.args.get("section", "todo")
        if section == "teams":
            teams = [t.strip() for t in request.form.getlist("team_names") if t.strip()]
            cfg["default_team_names"] = teams
            save_settings(cfg)
            flash("Standard-Teams gespeichert.", "success")
        else:
            cats = [c.strip() for c in request.form.getlist("categories") if c.strip()]
            cfg["default_todo_categories"] = cats
            save_settings(cfg)
            flash("Standard-Kategorien gespeichert.", "success")
        return redirect(url_for("sa_settings"))
    cats  = cfg.get("default_todo_categories", DEFAULT_TODO_CATEGORIES)
    teams = cfg.get("default_team_names", ["Team Abrechnung", "Team Personal"])
    return render_template("sa_settings.html", categories=cats, default_teams=teams)


FETCHER_FEED_OUTPUT = os.path.join(DATA_DIR, "fetcher_feed_latest.json")

@app.route("/admin/fetcher/feed/run", methods=["POST"])
@role_required("superadmin")
def fetcher_feed_run():
    from fetcher_feed import run
    try:
        result = run()
        return jsonify({"ok": True, "result": result})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500

@app.route("/admin/fetcher/feed/status")
@role_required("admin", "superadmin")
def fetcher_feed_status():
    if not os.path.exists(FETCHER_FEED_OUTPUT):
        return jsonify({"exists": False})
    with open(FETCHER_FEED_OUTPUT, encoding="utf-8") as f:
        data = json.load(f)
    return jsonify({
        "exists":            True,
        "run_at":            data.get("run_at", ""),
        "new_entries_total": data.get("new_entries_total", 0),
        "sources_checked":   data.get("sources_checked", 0),
        "sources_ok":        data.get("sources_ok", 0),
        "sources_error":     data.get("sources_error", 0),
    })

@app.route("/admin/fetcher/feed/export")
@role_required("admin", "superadmin")
def fetcher_feed_export():
    if not os.path.exists(FETCHER_FEED_OUTPUT):
        flash("Noch keine Fetcher-Ergebnisse. Bitte zuerst Fletcher starten.", "error")
        return redirect(url_for("agents"))
    with open(FETCHER_FEED_OUTPUT, encoding="utf-8") as f:
        data = json.load(f)
    from fetcher_feed import export_text
    content = export_text(data)
    date_str = data.get("run_at", "")[:10]
    return Response(
        content,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="fletcher_{date_str}.txt"'},
    )

FETCHER_SCRAPE_OUTPUT = os.path.join(DATA_DIR, "fetcher_scrape_latest.json")

@app.route("/admin/fetcher/scrape/run", methods=["POST"])
@role_required("superadmin")
def fetcher_scrape_run():
    from fetcher_scrape import run
    try:
        result = run()
        return jsonify({"ok": True, "result": result})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500

@app.route("/admin/fetcher/scrape/status")
@role_required("admin", "superadmin")
def fetcher_scrape_status():
    if not os.path.exists(FETCHER_SCRAPE_OUTPUT):
        return jsonify({"exists": False})
    with open(FETCHER_SCRAPE_OUTPUT, encoding="utf-8") as f:
        data = json.load(f)
    return jsonify({
        "exists": True,
        "run_at": data.get("run_at"),
        "new_entries_total": data.get("new_entries_total", 0),
        "sources_checked": data.get("sources_checked", 0),
        "sources_ok": data.get("sources_ok", 0),
        "sources_error": data.get("sources_error", 0),
    })

@app.route("/admin/fetcher/scrape/export")
@role_required("admin", "superadmin")
def fetcher_scrape_export():
    if not os.path.exists(FETCHER_SCRAPE_OUTPUT):
        flash("Noch keine Fetcher-Ergebnisse. Bitte zuerst Frank starten.", "error")
        return redirect(url_for("agents"))
    with open(FETCHER_SCRAPE_OUTPUT, encoding="utf-8") as f:
        data = json.load(f)
    from fetcher_scrape import export_text
    content = export_text(data)
    date_str = data.get("run_at", "")[:10]
    return Response(
        content,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="frank_{date_str}.txt"'},
    )

SELECTOR_STATUS_FILE = os.path.join(DATA_DIR, "selector_finder_status.json")
MAJA_STATUS_FILE     = os.path.join(DATA_DIR, "maja_status.json")
MAJA_HISTORY_FILE    = os.path.join(DATA_DIR, "maja_history.json")
MAJA_INSIGHTS_FILE   = os.path.join(DATA_DIR, "maja_insights.json")

# ── APScheduler ───────────────────────────────────────────────────────────────
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

_scheduler = BackgroundScheduler(timezone="Europe/Berlin")
_scheduler.start()

def _maja_scheduled_job():
    """Called by APScheduler on schedule."""
    import maja_runner
    cfg     = load_agent_configs()
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return
    maja_cfg = cfg.get("agents", {}).get("source_maintenance", {})
    model    = maja_cfg.get("maja_model", "claude-haiku-4-5-20251001")
    try:
        maja_runner.run(maja_cfg, api_key, model=model)
    except Exception:
        pass

def _apply_maja_schedule(cfg):
    """Read schedule config and register/replace the APScheduler job."""
    maja = cfg.get("agents", {}).get("source_maintenance", {})
    job_id = "maja_scheduled"
    _scheduler.remove_job(job_id) if _scheduler.get_job(job_id) else None
    if not maja.get("schedule_enabled"):
        return
    stype   = maja.get("schedule_type", "")
    s_time  = maja.get("schedule_time", "08:00")
    try:
        hour, minute = s_time.split(":")
    except Exception:
        hour, minute = "8", "0"
    if stype == "monthly":
        day = maja.get("schedule_day", "1")
        _scheduler.add_job(_maja_scheduled_job, CronTrigger(day=day, hour=hour, minute=minute),
                           id=job_id, replace_existing=True)
    elif stype == "weekly":
        weekday = maja.get("schedule_weekday", "mon")
        _scheduler.add_job(_maja_scheduled_job, CronTrigger(day_of_week=weekday, hour=hour, minute=minute),
                           id=job_id, replace_existing=True)

# Apply schedule on startup
with app.app_context():
    try:
        _apply_maja_schedule(load_agent_configs())
    except Exception:
        pass

@app.route("/admin/selector/run", methods=["POST"])
@role_required("superadmin")
def selector_run():
    import selector_finder
    cfg = load_agent_configs()
    sol = cfg.get("agents", {}).get("sol_selector", {})
    try:
        result = selector_finder.run(
            timeout=int(sol.get("timeout", 8)),
            min_links=int(sol.get("min_links", 3)),
            max_per_run=int(sol.get("max_per_run", 30)),
        )
        return jsonify({"ok": True, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/admin/selector/status")
@role_required("admin", "superadmin")
def selector_status():
    if not os.path.exists(SELECTOR_STATUS_FILE):
        return jsonify({"exists": False})
    with open(SELECTOR_STATUS_FILE, encoding="utf-8") as f:
        data = json.load(f)
    return jsonify({
        "exists":             True,
        "run_at":             data.get("run_at", ""),
        "endpoints_checked":  data.get("endpoints_checked", 0),
        "updated":            data.get("updated", 0),
        "failed":             data.get("failed", 0),
        "remaining":          data.get("remaining", 0),
    })

_maja_running = False

@app.route("/admin/maja/run", methods=["POST"])
@role_required("superadmin")
def maja_run():
    global _maja_running
    if _maja_running:
        return jsonify({"ok": False, "error": "Maja läuft bereits."})
    import maja_runner, threading
    cfg     = load_agent_configs()
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY nicht gesetzt."})
    maja_cfg = cfg.get("agents", {}).get("source_maintenance", {})
    model    = maja_cfg.get("maja_model", "claude-haiku-4-5-20251001")

    from_idx = request.json.get("from_idx") if request.json else None
    to_idx   = request.json.get("to_idx")   if request.json else None

    def _run():
        global _maja_running
        _maja_running = True
        try:
            maja_runner.run(maja_cfg, api_key, model=model, from_idx=from_idx, to_idx=to_idx)
        except Exception as e:
            import traceback, json as _json
            err = {"run_at": __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                   "error": str(e), "traceback": traceback.format_exc()}
            with open(MAJA_STATUS_FILE, "w", encoding="utf-8") as f:
                _json.dump(err, f, ensure_ascii=False, indent=2)
            app.logger.error("Maja thread error: %s", traceback.format_exc())
        finally:
            _maja_running = False

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "started": True})

@app.route("/admin/maja/test", methods=["POST"])
@role_required("superadmin")
def maja_test():
    import maja_runner
    cfg     = load_agent_configs()
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY nicht gesetzt."})
    maja_cfg = cfg.get("agents", {}).get("source_maintenance", {})
    model    = maja_cfg.get("maja_model", "claude-haiku-4-5-20251001")
    try:
        result = maja_runner.test_run(maja_cfg, api_key, model=model)
        return jsonify({"ok": True, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/admin/maja/status")
@role_required("admin", "superadmin")
def maja_status():
    job = _scheduler.get_job("maja_scheduled")
    next_run = job.next_run_time.strftime("%Y-%m-%d %H:%M") if job and job.next_run_time else None
    if _maja_running:
        return jsonify({"exists": False, "running": True, "next_run": next_run})
    if not os.path.exists(MAJA_STATUS_FILE):
        return jsonify({"exists": False, "running": False, "next_run": next_run})
    with open(MAJA_STATUS_FILE, encoding="utf-8") as f:
        data = json.load(f)
    return jsonify({
        "exists":          True,
        "running":         False,
        "run_at":          data.get("run_at", ""),
        "sources_total":   data.get("sources_total", 0),
        "batches":         data.get("batches", 0),
        "endpoints_added": data.get("endpoints_added", 0),
        "errors":          data.get("errors", []),
        "next_run":        next_run,
    })

@app.route("/admin/maja/schedule", methods=["POST"])
@role_required("superadmin")
def maja_schedule():
    cfg  = load_agent_configs()
    maja = cfg["agents"].setdefault("source_maintenance", {})
    body = request.get_json(silent=True) or {}
    maja["schedule_enabled"] = bool(body.get("schedule_enabled"))
    maja["schedule_type"]    = body.get("schedule_type", "monthly")
    maja["schedule_day"]     = body.get("schedule_day", "1")
    maja["schedule_weekday"] = body.get("schedule_weekday", "mon")
    maja["schedule_time"]    = body.get("schedule_time", "08:00")
    maja["maja_model"]       = body.get("maja_model", "claude-haiku-4-5-20251001")
    save_agent_configs(cfg)
    _apply_maja_schedule(cfg)
    job = _scheduler.get_job("maja_scheduled")
    next_run = job.next_run_time.strftime("%Y-%m-%d %H:%M") if job and job.next_run_time else None
    return jsonify({"ok": True, "next_run": next_run})

def _maja_quality(sources):
    total = len(sources)
    with_ep   = sum(1 for s in sources if s.get("endpoints"))
    no_ep     = total - with_ep
    by_type   = {"feed": 0, "scrape": 0, "pdf": 0, "search": 0}
    validated = 0  # scrape endpoints with a selector (Sol confirmed)
    for s in sources:
        for ep in s.get("endpoints", []):
            t = ep.get("agent", "")
            if t in by_type:
                by_type[t] += 1
            if t == "scrape" and ep.get("selector"):
                validated += 1
    return {
        "total":     total,
        "with_ep":   with_ep,
        "no_ep":     no_ep,
        "pct":       round(with_ep / total * 100, 1) if total else 0,
        "by_type":   by_type,
        "validated": validated,
    }

@app.route("/admin/maja/insights")
@role_required("admin", "superadmin")
def maja_insights():
    if not os.path.exists(MAJA_INSIGHTS_FILE):
        return jsonify([])
    with open(MAJA_INSIGHTS_FILE, encoding="utf-8") as f:
        return jsonify(json.load(f))

@app.route("/admin/maja/insights/<insight_id>", methods=["POST"])
@role_required("admin", "superadmin")
def maja_insight_update(insight_id):
    if not os.path.exists(MAJA_INSIGHTS_FILE):
        return jsonify({"ok": False})
    with open(MAJA_INSIGHTS_FILE, encoding="utf-8") as f:
        insights = json.load(f)
    status = request.json.get("status") if request.json else None
    if status not in ("neu", "geprüft", "abgelehnt"):
        return jsonify({"ok": False, "error": "Ungültiger Status"})
    for item in insights:
        if item.get("id") == insight_id:
            item["status"] = status
            item["status_updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            break
    with open(MAJA_INSIGHTS_FILE, "w", encoding="utf-8") as f:
        json.dump(insights, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True})

@app.route("/admin/maja/report")
@role_required("admin", "superadmin")
def maja_report():
    sources = load_sources()
    quality = _maja_quality(sources)
    history = []
    if os.path.exists(MAJA_HISTORY_FILE):
        with open(MAJA_HISTORY_FILE, encoding="utf-8") as f:
            history = json.load(f)[:10]
    return jsonify({"quality": quality, "history": history})

@app.route("/superadmin/backup/download")
@role_required("superadmin")
def backup_download():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Code files
        code_files = ["app.py", "requirements.txt", "init_users.py", ".env.example"]
        for fname in code_files:
            fpath = os.path.join(base_dir, fname)
            if os.path.exists(fpath):
                zf.write(fpath, os.path.join("code", fname))
        # Templates
        tmpl_dir = os.path.join(base_dir, "templates")
        if os.path.isdir(tmpl_dir):
            for fname in os.listdir(tmpl_dir):
                fpath = os.path.join(tmpl_dir, fname)
                if os.path.isfile(fpath):
                    zf.write(fpath, os.path.join("code", "templates", fname))
        # Static files
        static_dir = os.path.join(base_dir, "static")
        if os.path.isdir(static_dir):
            for root, _, files in os.walk(static_dir):
                for fname in files:
                    fpath = os.path.join(root, fname)
                    arcname = os.path.join("code", "static", os.path.relpath(fpath, static_dir))
                    zf.write(fpath, arcname)
        # Data files
        if os.path.isdir(DATA_DIR):
            for root, _, files in os.walk(DATA_DIR):
                for fname in files:
                    fpath = os.path.join(root, fname)
                    arcname = os.path.join("data", os.path.relpath(fpath, DATA_DIR))
                    zf.write(fpath, arcname)
    buf.seek(0)
    from flask import send_file
    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"branchenradar_backup_{ts}.zip"
    )


@app.route("/admin/signals/import", methods=["GET", "POST"])
@role_required("admin", "superadmin")
def signal_import():
    sections = load_sections()
    if request.method == "POST":
        raw = ""
        f = request.files.get("file")
        if f and f.filename:
            raw = f.read().decode("utf-8-sig")
        else:
            raw = request.form.get("csv_text", "").strip()
        if not raw:
            flash("Keine Daten eingegeben.", "error")
            return redirect(url_for("signal_import"))
        imported, errors = _parse_signal_csv(raw, sections)
        if imported:
            existing = load_signals()
            existing.extend(imported)
            save_signals(existing)
        msg_parts = []
        if imported:
            msg_parts.append(f"{len(imported)} Signal{'e' if len(imported)!=1 else ''} importiert")
        if errors:
            msg_parts.append(f"{len(errors)} Zeile{'n' if len(errors)!=1 else ''} übersprungen")
        flash((" · ".join(msg_parts)) or "Nichts importiert.", "success" if imported else "error")
        return redirect(url_for("signal_list"))
    return render_template("admin_signals_import.html",
        categories=CATEGORIES, status_labels=STATUS_LABELS, sections=sections)


def _parse_signal_csv(raw, sections):
    """Parse CSV or TSV text into signal dicts. Returns (imported, errors)."""
    # Detect separator: tab wins if present on first line, then semicolon, then comma
    first_line = raw.split("\n")[0]
    if "\t" in first_line:
        sep = "\t"
    elif ";" in first_line:
        sep = ";"
    else:
        sep = ","

    # Build lookup maps
    status_map = {
        "radar": "radar", "im radar": "radar",
        "develop": "develop", "in entwicklung": "develop", "entwicklung": "develop",
        "announced": "announced", "angekündigt": "announced", "angekundigt": "announced",
        "active": "active", "verfügbar": "active", "in kraft": "active", "verfugbar": "active",
        "action": "action", "handlungsbedarf": "action",
    }
    cat_map = {
        "krankenkassen": "krankenkassen", "krankenkassen & gkv": "krankenkassen", "gkv": "krankenkassen",
        "digital": "digital", "digitalisierung": "digital", "digitalisierung & ti": "digital", "ti": "digital",
        "gesetze": "gesetze", "regulatorien": "gesetze", "gesetze & regulatorien": "gesetze",
        "personal": "personal", "tarife": "personal", "personal & tarife": "personal",
        "praxis": "praxis", "praxismanagement": "praxis",
    }
    sec_by_name = {s["name"].lower(): s["id"] for s in sections}
    sec_by_id   = {s["id"]: s["id"] for s in sections}

    # Header aliases → field name
    col_map = {
        "titel": "title", "title": "title",
        "zusammenfassung": "summary", "summary": "summary",
        "detail": "detail", "detailtext": "detail",
        "status": "status",
        "kategorie": "category", "category": "category",
        "fachbereich": "section_ids", "fachbereiche": "section_ids", "section_ids": "section_ids", "sektionen": "section_ids",
        "quelle": "source", "source": "source",
        "quelle-url": "source_url", "quellenurl": "source_url", "source_url": "source_url", "url": "source_url",
        "datum": "date", "date": "date",
        "veröffentlicht am": "published_at", "published_at": "published_at",
        "beschlossen am": "decision_at", "decision_at": "decision_at",
        "im bundesanzeiger": "gazetted_at", "gazetted_at": "gazetted_at",
        "gilt ab": "effective_from", "effective_from": "effective_from",
        "frist bis": "deadline_at", "deadline_at": "deadline_at",
        "frist": "deadline", "deadline": "deadline",
        "entwicklungsstand": "entwicklungsstand",
        "agent": "agent",
        "region": "region",
    }
    entwicklungsstand_map = {
        "beobachtung": "beobachtung",
        "entwurf": "entwurf",
        "konsultation": "konsultation",
        "beschlossen": "beschlossen",
        "veröffentlicht": "veroeffentlicht", "veroeffentlicht": "veroeffentlicht",
        "verkündet": "veroeffentlicht", "verkundet": "veroeffentlicht",
        "verkündet / veröffentlicht": "veroeffentlicht",
        "in kraft ab": "in_kraft_kuenftig", "in_kraft_kuenftig": "in_kraft_kuenftig",
        "in kraft": "in_kraft", "in_kraft": "in_kraft",
        "aufgehoben": "aufgehoben", "abgelöst": "aufgehoben", "abgeloest": "aufgehoben",
        "aufgehoben / abgelöst": "aufgehoben",
    }

    reader = csv.DictReader(io.StringIO(raw), delimiter=sep)
    # Normalise header names
    if reader.fieldnames is None:
        return [], ["Keine Spaltenüberschriften erkannt"]
    headers = {h: col_map.get(h.strip().lower().replace(" ", "-"), None) for h in reader.fieldnames}

    imported, errors = [], []
    for i, row in enumerate(reader, start=2):
        mapped = {}
        for raw_col, field in headers.items():
            if field and raw_col in row:
                mapped[field] = (row[raw_col] or "").strip()

        title = mapped.get("title", "")
        if not title:
            errors.append(f"Zeile {i}: kein Titel")
            continue

        # Resolve status
        raw_status = mapped.get("status", "radar").lower().strip()
        status = status_map.get(raw_status, "radar")

        # Resolve category
        raw_cat = mapped.get("category", "").lower().strip()
        category = cat_map.get(raw_cat, "praxis")

        # Resolve section_ids
        raw_secs = mapped.get("section_ids", "")
        sec_ids = []
        for part in raw_secs.replace(";", ",").split(","):
            part = part.strip()
            if not part:
                continue
            sid = sec_by_id.get(part) or sec_by_name.get(part.lower())
            if sid:
                sec_ids.append(sid)

        sig = {
            "id":         next_signal_raw_id(),
            "agent":      mapped.get("agent", ""),
            "region":     mapped.get("region", "Bundesweit"),
            "title":      title,
            "summary":    mapped.get("summary", ""),
            "detail":     mapped.get("detail", ""),
            "status":     status,
            "category":   category,
            "source":     mapped.get("source", ""),
            "source_url": mapped.get("source_url", ""),
            "entwicklungsstand": entwicklungsstand_map.get(mapped.get("entwicklungsstand", "").lower().strip(), mapped.get("entwicklungsstand", "")),
            "date":         mapped.get("date", datetime.now().strftime("%Y-%m-%d")),
            "published_at": mapped.get("published_at", ""),
            "decision_at":  mapped.get("decision_at", ""),
            "gazetted_at":  mapped.get("gazetted_at", ""),
            "effective_from": mapped.get("effective_from", ""),
            "deadline_at":  mapped.get("deadline_at", ""),
            "deadline":     mapped.get("deadline", ""),
            "section_ids":  sec_ids,
            "first_seen_at": datetime.now().isoformat(),
            "created_at":   datetime.now().isoformat(),
        }
        imported.append(sig)
    return imported, errors


# ── Helpers ───────────────────────────────────────────────────────────────────

def _create_user(form, redirect_to, force_entity=None, force_role=None):
    username = form.get("username", "").strip()
    password = form.get("password", "").strip()
    name     = form.get("name", "").strip()
    role     = force_role or form.get("role", "user")
    entity_id = force_entity or form.get("entity_id") or None

    if not username or not password:
        flash("Benutzername und Passwort sind erforderlich.", "error")
        return redirect(redirect_to)
    users = load_users()
    if any(u["username"] == username for u in users):
        flash(f"Benutzername '{username}' existiert bereits.", "error")
        return redirect(redirect_to)
    users.append({
        "id":        str(uuid.uuid4()),
        "username":  username,
        "pass_hash": _hash(password),
        "role":      role,
        "entity_id": entity_id,
        "name":      name or username,
    })
    save_users(users)
    flash(f"Nutzer '{username}' angelegt.", "success")
    return redirect(redirect_to)

def _signal_from_form(form, existing_id=None):
    sources = []
    for i in range(1, 4):
        name = form.get(f"source_{i}_name", "").strip()
        url  = form.get(f"source_{i}_url",  "").strip()
        date = form.get(f"source_{i}_date", "").strip()
        if name or url:
            sources.append({"name": name, "url": url, "date": date})
    primary_date = sources[0]["date"] if sources and sources[0].get("date") else datetime.now().strftime("%Y-%m-%d")
    return {
        "id":                existing_id or str(uuid.uuid4()),
        "title":             form.get("title", "").strip(),
        "summary":           form.get("summary", "").strip(),
        "detail":            form.get("detail", "").strip(),
        "category":          form.get("category", "gesetze"),
        "priority":          form.get("priority", "sollte"),
        "entwicklungsstand": form.get("entwicklungsstand", "beobachtung"),
        "handlungszeitpunkt":form.get("handlungszeitpunkt", ""),
        "naechster_schritt": form.get("naechster_schritt", "").strip(),
        "betroffene_rollen": form.getlist("betroffene_rollen"),
        "aufwand":           form.get("aufwand", ""),
        "themen_id":         form.get("themen_id", "").strip(),
        "sources":           sources,
        "date":              primary_date,
        "published_at":      form.get("published_at", "").strip() or None,
        "decision_at":       form.get("decision_at", "").strip() or None,
        "gazetted_at":       form.get("gazetted_at", "").strip() or None,
        "effective_from":    form.get("effective_from", "").strip() or None,
        "deadline_at":       form.get("deadline_at", "").strip() or None,
        "deadline":          form.get("deadline", "").strip() or None,
        "section_ids":       form.getlist("section_ids"),
        "reporting_status":  form.get("reporting_status", "").strip(),
        "agent":             form.get("agent", "").strip(),
        "region":            form.get("region", "Bundesweit").strip(),
    }


# ── Signal MATCH / FINAL ─────────────────────────────────────────────────────

@app.route("/admin/signals/match", methods=["GET", "POST"])
@role_required("admin", "superadmin")
def signal_match():
    matches = load_signal_matches()
    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "add":
            row = {
                "id": str(uuid.uuid4()),
                "sig1": request.form.get("sig1", "").strip(),
                "pairs": []
            }
            for i in range(2, 6):
                sig_id = request.form.get(f"sig{i}", "").strip()
                sig_type = request.form.get(f"sig{i}_type", "").strip()
                if sig_id:
                    row["pairs"].append({"id": sig_id, "type": sig_type})
            matches.append(row)
            save_signal_matches(matches)
            flash("Gruppe gespeichert.", "success")
        elif action == "delete":
            row_id = request.form.get("row_id", "")
            matches = [m for m in matches if m["id"] != row_id]
            save_signal_matches(matches)
            flash("Gruppe gelöscht.", "success")
        return redirect(url_for("signal_match"))
    signals = load_signals()
    sig_map = {s["id"]: s.get("title", s["id"]) for s in signals}
    return render_template("admin_signal_matches.html", matches=matches, sig_map=sig_map)

@app.route("/admin/signals/final", methods=["GET", "POST"])
@role_required("admin", "superadmin")
def signal_final():
    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "import_csv":
            raw = request.form.get("csv_data", "").strip()
            if raw:
                sections = load_sections()
                imported, errors = _parse_signal_csv(raw, sections)
                if imported:
                    archive_signal_final()
                    ts = datetime.now().strftime("%y%m%d%H%M")
                    save_signal_final(imported, ts)
                    flash(f"{len(imported)} Signale als neue FINAL-Version gespeichert (Timestamp: {ts}).", "success")
                if errors:
                    for e in errors[:5]:
                        flash(e, "error")
        elif action == "delete":
            sig_id = request.form.get("sig_id", "")
            data = load_signal_final()
            data["signals"] = [s for s in data["signals"] if s["id"] != sig_id]
            save_signal_final(data["signals"], data["timestamp"])
            flash("Signal gelöscht.", "success")
        return redirect(url_for("signal_final"))
    data = load_signal_final()
    return render_template("admin_signal_final.html",
        signals=data["signals"], timestamp=data["timestamp"],
        categories=CATEGORIES, priority_labels=SIGNAL_PRIORITY_LABELS,
        entwicklungsstand_labels=ENTWICKLUNGSSTAND_LABELS)

@app.route("/admin/signals/archive")
@role_required("admin", "superadmin")
def signal_archive():
    archives = list_signal_final_archives()
    return render_template("admin_signal_archive.html", archives=archives)

@app.route("/admin/signals/archive/<ts>")
@role_required("admin", "superadmin")
def signal_archive_view(ts):
    data = load_signal_final_archive(ts)
    if not data:
        abort(404)
    signals = data.get("signals", []) if isinstance(data, dict) else data
    timestamp = data.get("timestamp", ts) if isinstance(data, dict) else ts
    return render_template("admin_signal_archive_view.html",
        signals=signals, timestamp=timestamp, ts=ts,
        categories=CATEGORIES, priority_labels=SIGNAL_PRIORITY_LABELS)


# ── Signal Excel-Exports ─────────────────────────────────────────────────────

def _signals_to_xlsx(signals, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "Signale"
    headers = [
        "ID", "Titel", "Zusammenfassung", "Detail",
        "Kategorie", "Priorität", "Entwicklungsstand", "Reporting-Status",
        "Datum", "Frist",
        "Quelle 1", "Quellenlink 1", "Datum Q1",
        "Quelle 2", "Quellenlink 2", "Datum Q2",
        "Quelle 3", "Quellenlink 3", "Datum Q3",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1D3D28")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False)
    prio_map = {"muss": "MUSS", "sollte": "SOLLTE", "kann": "KANN"}
    for s in signals:
        src = s.get("sources", [])
        def src_f(i, field): return src[i].get(field, "") if len(src) > i else ""
        ws.append([
            s.get("id", ""),
            s.get("title", ""),
            s.get("summary", ""),
            s.get("detail", ""),
            s.get("category", ""),
            prio_map.get(s.get("priority", ""), s.get("priority", "")),
            s.get("entwicklungsstand", ""),
            s.get("reporting_status", ""),
            s.get("date", ""),
            s.get("deadline", "") or "",
            src_f(0, "name"), src_f(0, "url"), src_f(0, "date"),
            src_f(1, "name"), src_f(1, "url"), src_f(1, "date"),
            src_f(2, "name"), src_f(2, "url"), src_f(2, "date"),
        ])
    col_widths = [14, 40, 60, 60, 14, 10, 18, 14, 12, 12, 20, 40, 12, 20, 40, 12, 20, 40, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)

@app.route("/admin/signals/export/match")
@role_required("admin", "superadmin")
def signal_export_match():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io
    matches = load_signal_matches()
    wb = Workbook()
    ws = wb.active
    ws.title = "MATCH"
    headers = ["sig1", "sig2", "sig2_type", "sig3", "sig3_type", "sig4", "sig4_type", "sig5", "sig5_type"]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="1D3D28")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont
    for row in matches:
        r = [row.get("sig1", "")]
        for p in (row.get("pairs", []) + [{}, {}, {}, {}])[:4]:
            r += [p.get("id", ""), p.get("type", "")]
        ws.append(r)
    for i, w in enumerate([16, 16, 6, 16, 6, 16, 6, 16, 6], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"signale_match_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.route("/admin/signals/export/raw")
@role_required("admin", "superadmin")
def signal_export_raw():
    signals = load_signals()
    signals.sort(key=lambda s: s.get("id", ""))
    return _signals_to_xlsx(signals, f"signale_raw_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.route("/admin/signals/export/final")
@role_required("admin", "superadmin")
def signal_export_final():
    data = load_signal_final()
    signals = sorted(data["signals"], key=lambda s: s.get("id", ""))
    return _signals_to_xlsx(signals, f"signale_final_{datetime.now().strftime('%Y%m%d')}.xlsx")


@app.route("/admin/sources/export")
@role_required("admin", "superadmin")
def sources_export():
    from flask import send_file
    import openpyxl
    sources_list = load_sources()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quellen"
    headers = [
        "Kürzel", "Name", "URL", "Region", "Endpoints", "Primärkategorie", "Relevante Rollen",
        "Agenten", "Priorität", "Zugang", "Aktualisierung", "Status", "Hinweise",
        "Agenten-Hinweis (scrape)", "Agenten-Hinweis (feed)",
        "Agenten-Hinweis (pdf)", "Agenten-Hinweis (search)",
        "Datum Hinzugefügt", "Datum Letzte Änderung", "Kommentar",
    ]
    ws.append(headers)
    for s in sorted(sources_list, key=lambda x: x.get("kuerzel", "")):
        ah = {h.get("agent", ""): h.get("text", "") for h in s.get("agent_hints", [])}
        eps = s.get("endpoints", [])
        eps_str = " | ".join(
            f"{e.get('agent','')}:{e.get('url','')} ({e.get('label','')})" for e in eps
        ) if eps else ""
        ws.append([
            s.get("kuerzel", ""),
            s.get("name", ""),
            s.get("url", ""),
            s.get("region", "Bundesweit"),
            eps_str,
            s.get("primary_category", ""),
            ", ".join(s.get("relevant_roles", [])),
            ", ".join(s.get("ingestion_methods", [])),
            s.get("priority", ""),
            s.get("zugang", ""),
            s.get("expected_update", ""),
            s.get("status", ""),
            s.get("notes", ""),
            ah.get("scrape", ""),
            ah.get("feed", ""),
            ah.get("pdf", ""),
            ah.get("search", ""),
            s.get("created_at", ""),
            s.get("updated_at", ""),
            s.get("comment", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"quellen_{datetime.now().strftime('%Y%m%d')}.xlsx")


def _parse_sources_tsv(raw_text):
    """Parse TSV from agent output or app export into source dicts."""
    lines = [l for l in raw_text.strip().splitlines() if l.strip()]
    if not lines:
        return []
    # Detect header row
    header = [h.strip().lower() for h in lines[0].split("\t")]
    col = {
        "kuerzel":        next((i for i, h in enumerate(header) if "kürzel" in h or "kurzel" in h), None),
        "name":           next((i for i, h in enumerate(header) if h == "name"), None),
        "url":            next((i for i, h in enumerate(header) if h in ("url", "url (konkret)")), None),
        "region":         next((i for i, h in enumerate(header) if h == "region"), None),
        "endpoints":      next((i for i, h in enumerate(header) if "endpoint" in h), None),
        "primary_category": next((i for i, h in enumerate(header) if "primär" in h or "kategorie" in h), None),
        "relevant_roles": next((i for i, h in enumerate(header) if "rollen" in h), None),
        "ingestion_methods": next((i for i, h in enumerate(header) if "agenten" == h or h in ("agenten", "empfohlener agent")), None),
        "priority":       next((i for i, h in enumerate(header) if "priorität" in h or "prioritat" in h), None),
        "zugang":         next((i for i, h in enumerate(header) if "zugang" in h), None),
        "expected_update": next((i for i, h in enumerate(header) if "aktualis" in h), None),
        "status":         next((i for i, h in enumerate(header) if h == "status"), None),
        "notes":          next((i for i, h in enumerate(header) if "hinweise" in h and "agenten" not in h), None),
        "hint_scrape":    next((i for i, h in enumerate(header) if "scrape" in h), None),
        "hint_feed":      next((i for i, h in enumerate(header) if "feed" in h and "url" not in h), None),
        "hint_pdf":       next((i for i, h in enumerate(header) if "pdf" in h), None),
        "hint_search":    next((i for i, h in enumerate(header) if "search" in h or "suche" in h), None),
        "created_at":     next((i for i, h in enumerate(header) if "hinzugefügt" in h or "hinzugefugt" in h), None),
        "updated_at":     next((i for i, h in enumerate(header) if "letzte" in h and "änderung" in h), None),
        "comment":        next((i for i, h in enumerate(header) if "kommentar" in h), None),
    }
    def cell(row, key):
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return ""
        return row[idx].strip()
    results = []
    for line in lines[1:]:
        parts = line.split("\t")
        name = cell(parts, "name")
        if not name:
            continue
        agent_hints = []
        for agent_key in ("hint_scrape", "hint_feed", "hint_pdf", "hint_search"):
            text = cell(parts, agent_key)
            if text:
                agent_hints.append({"agent": agent_key.replace("hint_", ""), "text": text})
        methods_raw = cell(parts, "ingestion_methods")
        methods = [m.strip() for m in methods_raw.replace(",", " ").split() if m.strip() in ("scrape", "feed", "pdf", "search")]
        roles_raw = cell(parts, "relevant_roles")
        roles = [r.strip() for r in roles_raw.split(",") if r.strip()]
        # Parse endpoints from "agent:url (label) | ..." format
        eps_raw = cell(parts, "endpoints")
        endpoints = []
        if eps_raw:
            for ep_str in eps_raw.split(" | "):
                ep_str = ep_str.strip()
                if not ep_str:
                    continue
                label = ""
                if "(" in ep_str and ep_str.endswith(")"):
                    ep_str, label = ep_str[:-1].rsplit("(", 1)
                    label = label.strip()
                    ep_str = ep_str.strip()
                selector = ""
                if "[" in ep_str and ep_str.endswith("]"):
                    ep_str, selector = ep_str[:-1].rsplit("[", 1)
                    selector = selector.strip()
                    ep_str = ep_str.strip()
                if ":" in ep_str:
                    agent_key, url = ep_str.split(":", 1)
                else:
                    agent_key, url = "", ep_str
                if url.strip():
                    ep = {"url": url.strip(), "agent": agent_key.strip(), "label": label}
                    if selector:
                        ep["selector"] = selector
                    endpoints.append(ep)
        results.append({
            "kuerzel":           cell(parts, "kuerzel"),
            "name":              name,
            "url":               cell(parts, "url"),
            "region":            cell(parts, "region") or "Bundesweit",
            "endpoints":         endpoints,
            "primary_category":  cell(parts, "primary_category"),
            "relevant_roles":    roles,
            "ingestion_methods": methods,
            "priority":          cell(parts, "priority") or "mittel",
            "zugang":            cell(parts, "zugang"),
            "expected_update":   cell(parts, "expected_update"),
            "status":            cell(parts, "status"),
            "notes":             cell(parts, "notes"),
            "agent_hints":       agent_hints,
            "created_at":        cell(parts, "created_at"),
            "updated_at":        cell(parts, "updated_at"),
            "comment":           cell(parts, "comment"),
        })
    return results


@app.route("/admin/sources/bulk-import", methods=["POST"])
@role_required("superadmin")
def sources_bulk_import():
    action = request.form.get("action", "preview")
    raw = request.form.get("tsv_data", "").strip()
    if not raw:
        flash("Keine Daten eingefügt.", "error")
        return redirect(url_for("sources"))
    incoming = _parse_sources_tsv(raw)
    if not incoming:
        flash("Keine gültigen Zeilen erkannt. Prüfe das Format.", "error")
        return redirect(url_for("sources"))
    existing = load_sources()
    existing_by_kuerzel = {s.get("kuerzel", "").upper(): s for s in existing if s.get("kuerzel")}
    today = datetime.now().strftime("%Y-%m-%d")
    if action == "preview":
        new_srcs, modified, unchanged = [], [], []
        for s in incoming:
            k = s.get("kuerzel", "").upper()
            if not k or k not in existing_by_kuerzel:
                new_srcs.append(s.get("name", k))
            else:
                ex = existing_by_kuerzel[k]
                changed_fields = []
                for field in ("url", "feed_url", "primary_category", "priority", "status", "notes",
                              "zugang", "expected_update", "ingestion_methods", "relevant_roles"):
                    if s.get(field) != ex.get(field):
                        changed_fields.append(field)
                if changed_fields:
                    modified.append({"name": s.get("name", k), "fields": changed_fields})
                else:
                    unchanged.append(s.get("name", k))
        return jsonify({
            "ok": True,
            "total": len(incoming),
            "new": new_srcs,
            "modified": modified,
            "unchanged": len(unchanged),
        })
    elif action == "confirm":
        merged = []
        for s in incoming:
            k = s.get("kuerzel", "").upper()
            ex = existing_by_kuerzel.get(k)
            src = dict(s)
            src["id"] = ex["id"] if ex else str(uuid.uuid4())
            # Preserve created_at from existing if not set
            if not src.get("created_at"):
                src["created_at"] = ex.get("created_at", today) if ex else today
            # Set updated_at to today if fields changed or agent set it
            if not src.get("updated_at") and ex:
                for field in ("url", "feed_url", "primary_category", "priority", "status",
                              "notes", "zugang", "expected_update", "ingestion_methods", "relevant_roles"):
                    if src.get(field) != ex.get(field):
                        src["updated_at"] = today
                        break
            elif not src.get("updated_at") and not ex:
                src["created_at"] = today
            merged.append(src)
        save_sources(merged)
        flash(f"Import abgeschlossen: {len(merged)} Quellen gespeichert.", "success")
        return jsonify({"ok": True, "redirect": url_for("sources")})


# ── Agent Reports ─────────────────────────────────────────────────────────────

@app.route("/admin/agent-reports/raw", methods=["GET", "POST"])
@role_required("admin", "superadmin")
def agent_reports_raw():
    reports = load_agent_reports()
    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "import":
            raw = request.form.get("tsv_data", "").strip()
            imported = 0
            for line in raw.splitlines():
                parts = line.split("\t")
                if len(parts) < 2:
                    continue
                source_name   = parts[0].strip()
                signal_count_s = parts[1].strip()
                if not source_name or source_name.lower() in ("quelle", "source"):
                    continue  # header row
                try:
                    signal_count = int(signal_count_s)
                except ValueError:
                    signal_count = 0
                issues    = parts[2].strip() if len(parts) > 2 else ""
                timestamp = parts[3].strip() if len(parts) > 3 else datetime.now().strftime("%y%m%d%H%M")
                agent_name = parts[4].strip() if len(parts) > 4 else ""
                reports.append({
                    "id": str(uuid.uuid4()),
                    "source_name": source_name,
                    "signal_count": signal_count,
                    "issues": issues,
                    "timestamp": timestamp,
                    "agent_name": agent_name,
                })
                imported += 1
            save_agent_reports(reports)
            flash(f"{imported} Einträge importiert.", "success")
        elif action == "delete_selected":
            ids_to_delete = set(request.form.getlist("delete_ids"))
            reports = [r for r in reports if r["id"] not in ids_to_delete]
            save_agent_reports(reports)
            flash(f"{len(ids_to_delete)} Einträge gelöscht.", "success")
        return redirect(url_for("agent_reports_raw"))
    reports_sorted = sorted(reports, key=lambda r: r.get("timestamp", ""), reverse=True)
    return render_template("admin_agent_reports_raw.html", reports=reports_sorted)

@app.route("/admin/agent-reports")
@role_required("admin", "superadmin")
def agent_reports():
    reports = load_agent_reports()
    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")
    # Filter by timestamp (JJMMDD prefix comparison)
    def ts_to_date(ts):
        if len(ts) >= 6:
            try:
                return datetime.strptime("20" + ts[:6], "%Y%m%d").date()
            except ValueError:
                pass
        return None
    if date_from or date_to:
        try:
            df = datetime.strptime(date_from, "%Y-%m-%d").date() if date_from else None
        except ValueError:
            df = None
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d").date() if date_to else None
        except ValueError:
            dt = None
        filtered = []
        for r in reports:
            rd = ts_to_date(r.get("timestamp", ""))
            if rd is None:
                filtered.append(r)
                continue
            if df and rd < df:
                continue
            if dt and rd > dt:
                continue
            filtered.append(r)
        reports = filtered
    # Build stats per agent
    SEARCH_AGENTS = [("scrape", "Scrape-Agent"), ("feed", "Feed-Agent"),
                     ("pdf", "PDF-Agent"), ("search", "Search-Agent")]
    stats = {}
    for key, label in SEARCH_AGENTS:
        agent_rows = [r for r in reports if r.get("agent_name") == key]
        by_source = {}
        for r in agent_rows:
            sn = r["source_name"]
            if sn not in by_source:
                by_source[sn] = {"count": 0, "issues": [], "timestamps": []}
            by_source[sn]["count"] += r.get("signal_count", 0)
            if r.get("issues"):
                by_source[sn]["issues"].append(r["issues"])
            by_source[sn]["timestamps"].append(r.get("timestamp", ""))
        high, mid, low, zero, error = [], [], [], [], []
        for sn, data in by_source.items():
            entry = {"source": sn, "count": data["count"],
                     "issues": "; ".join(set(data["issues"])),
                     "timestamps": data["timestamps"]}
            if data["issues"]:
                error.append(entry)
            elif data["count"] >= 5:
                high.append(entry)
            elif data["count"] >= 2:
                mid.append(entry)
            elif data["count"] == 1:
                low.append(entry)
            else:
                zero.append(entry)
        stats[key] = {
            "label": label,
            "high": sorted(high, key=lambda x: -x["count"]),
            "mid":  sorted(mid,  key=lambda x: -x["count"]),
            "low":  low,
            "zero": zero,
            "error": error,
        }
    return render_template("admin_agent_reports.html",
        stats=stats, date_from=date_from, date_to=date_to,
        search_agents=SEARCH_AGENTS)


# ── Agents ────────────────────────────────────────────────────────────────────

@app.route("/admin/agents", methods=["GET"])
@role_required("admin", "superadmin")
def agents():
    cfg          = load_agent_configs()
    sources_list = load_sources()
    settings_cfg = load_settings()
    hints        = settings_cfg.get("source_prompt_hints", "")
    src_counts   = {}
    for key, *_ in [("scrape",), ("feed",), ("pdf",), ("search",)]:
        src_counts[key] = sum(1 for s in sources_list if key in s.get("ingestion_methods", []))
    return render_template("agents.html",
        agent_configs=cfg, sources=sources_list, hints=hints,
        src_counts=src_counts)

@app.route("/admin/agents/save", methods=["POST"])
@role_required("superadmin")
def agents_save():
    cfg = load_agent_configs()
    field = request.form.get("field", "")
    agent_key = request.form.get("agent_key", "")
    value = request.form.get("value", "").strip()
    if field == "output_format":
        cfg["output_format"] = value
    elif field in ("persona", "method", "hint", "beschreibung", "konfiguration") and agent_key in cfg["agents"]:
        cfg["agents"][agent_key][field] = value
    elif field in ("date_from", "date_to", "last_date_from", "last_date_to") and agent_key in cfg["agents"]:
        cfg["agents"][agent_key][field] = value
    save_agent_configs(cfg)
    return ("", 204)


# ── Sources ───────────────────────────────────────────────────────────────────

REGION_OPTIONS = [
    "Bundesweit",
    "Baden-Württemberg", "Bayern", "Berlin", "Brandenburg", "Bremen",
    "Hamburg", "Hessen", "Mecklenburg-Vorpommern", "Niedersachsen",
    "Nordrhein-Westfalen", "Rheinland-Pfalz", "Saarland", "Sachsen",
    "Sachsen-Anhalt", "Schleswig-Holstein", "Thüringen",
    "EU / International",
]

SOURCE_PRIMARY_CATEGORIES = [
    "Abrechnung & Honorar",
    "Arzneimittel & Sicherheit",
    "Berufsrecht & Standespolitik",
    "DMP & Chronikerversorgung",
    "Diagnostik & Leitlinien",
    "Digitalisierung & TI",
    "EU-Regulatorik",
    "Evidenz & Nutzenbewertung",
    "Fachliches & Leitlinien",
    "Fachmedien & Trends",
    "GKV, Erstattung & Verträge",
    "Gesetzgebung & Politik",
    "Hygiene & Arbeitsschutz",
    "IT-Sicherheit & Datenschutz",
    "Impfungen & Prävention",
    "Infektionsschutz & Public Health",
    "International & Trends",
    "Markt & Produkte",
    "Medizinprodukte",
    "PKV, Gebühren & Erstattung",
    "Praxisbetrieb & Selbstverwaltung",
    "Praxismanagement",
    "Public Health & Prävention",
    "Qualität & Patientensicherheit",
    "Regionalrecht & Vollzug",
    "Regulatorik & Richtlinien",
    "Röntgen & Strahlenschutz",
    "Verordnung & Arzneimittel",
    "Versorgungsforschung & Daten",
    "Wissenschaft & Studien",
    "Sonstiges",
]
SOURCE_SECONDARY_CATEGORIES = [
    "Abrechnung", "Arzneimittel", "Bekanntmachungen", "Berichte", "Berufsrecht",
    "Berufe", "Chroniker", "Datenschutz", "Diagnostik", "Digitalisierung",
    "DMP", "EU-Recht", "Fachpresse", "GKV", "Gesetz", "Gesetzgebungsverfahren",
    "GOÄ", "GOZ", "HzV", "Hygiene", "Impfungen", "IT-Sicherheit",
    "KV-Recht", "Leitlinien", "Markt", "Medizinprodukte", "Nutzenbewertung",
    "PKV", "Prävention", "Praxismanagement", "Public Health", "Qualität",
    "Röntgen", "Selbstverwaltung", "Sozialrecht", "Stellungnahmen",
    "Strategien", "TI", "Tarifrecht", "Vergütung", "Verordnungen",
    "Versorgung", "Wissenschaft", "Zulassung",
]
SOURCE_RELEVANT_ROLES = [
    "Abrechnung", "IT", "MFA", "MFA/ZFA", "Praxisinhaber",
    "Praxismanagement", "QM/Hygiene", "ZFA", "ZMV/Abrechnung",
    "Zahnärztlicher Dienst", "Ärztlicher Dienst", "Ärztlicher/Zahnärztlicher Dienst",
]
SOURCE_ZUGANG_OPTIONS = [
    "öffentlich",
    "öffentlich/teilweise Login",
    "öffentlich/teilweise Mitgliederportal",
    "öffentlich/teilweise Paywall",
    "öffentlich/teilweise Volltext",
    "öffentlich/teilweise Lizenz",
    "öffentlich/Interaktion Login",
    "teilweise Paywall",
    "kommerziell",
    "Login/kommerziell",
]
SOURCE_UPDATE_OPTIONS = [
    "laufend", "täglich", "täglich/wöchentlich", "wöchentlich",
    "laufend/änderungsbezogen", "monatlich/änderungsbezogen", "monatlich",
    "quartalsweise/laufend", "quartalsweise", "änderungsbezogen",
    "ereignisbezogen", "jährlich",
]
SOURCE_INGESTION_METHODS = [
    ("scrape",  "Scrape",       "HTML-Seiten automatisiert abrufen und auslesen"),
    ("feed",    "Feed / API",   "RSS-, Atom- oder API-Endpunkte abonnieren"),
    ("pdf",     "PDF",          "Dokumente herunterladen und extrahieren"),
    ("search",  "Suche / Login","Suche, Login-Bereich oder Partnerschaftszugang"),
]
SOURCE_PRIORITIES = ["hoch", "mittel", "niedrig"]
SOURCE_STATUSES   = ["geplant", "in_recherche", "aktiv", "inaktiv"]
SOURCE_STATUS_LABELS = {
    "geplant":       "Geplant",
    "in_recherche":  "In Recherche",
    "aktiv":         "Aktiv",
    "inaktiv":       "Inaktiv",
}

@app.route("/admin/sources", methods=["GET", "POST"])
@role_required("admin", "superadmin")
def sources():
    cfg = load_settings()
    if request.method == "POST":
        action = request.form.get("action", "")
        sources_list = load_sources()

        def _endpoints_from_form(form):
            endpoints = []
            for i in range(1, 11):
                url      = form.get(f"ep_{i}_url", "").strip()
                agent    = form.get(f"ep_{i}_agent", "").strip()
                label    = form.get(f"ep_{i}_label", "").strip()
                selector = form.get(f"ep_{i}_selector", "").strip()
                if url:
                    ep = {"url": url, "agent": agent, "label": label}
                    if selector:
                        ep["selector"] = selector
                    endpoints.append(ep)
            return endpoints

        if action == "new":
            src = {
                "id":                   str(uuid.uuid4()),
                "kuerzel":              request.form.get("kuerzel", "").strip(),
                "name":                 request.form.get("name", "").strip(),
                "url":                  request.form.get("url", "").strip(),
                "region":               request.form.get("region", "Bundesweit").strip(),
                "endpoints":            _endpoints_from_form(request.form),
                "primary_category":     request.form.get("primary_category", "Sonstiges"),
                "secondary_categories": request.form.getlist("secondary_categories"),
                "relevant_roles":       request.form.getlist("relevant_roles"),
                "ingestion_methods":    request.form.getlist("ingestion_methods"),
                "priority":             request.form.get("priority", "mittel"),
                "status":               request.form.get("status", "").strip(),
                "zugang":               request.form.get("zugang", "").strip(),
                "expected_update":      request.form.get("expected_update", "").strip(),
                "notes":                request.form.get("notes", "").strip(),
                "created_at":           datetime.now().strftime("%Y-%m-%d"),
            }
            sources_list.append(src)
            save_sources(sources_list)
            flash("Quelle angelegt.", "success")

        elif action == "edit":
            sid = request.form.get("source_id", "")
            for s in sources_list:
                if s["id"] == sid:
                    s["kuerzel"]              = request.form.get("kuerzel", "").strip()
                    s["name"]                 = request.form.get("name", "").strip()
                    s["url"]                  = request.form.get("url", "").strip()
                    s["region"]               = request.form.get("region", "Bundesweit").strip()
                    s["endpoints"]            = _endpoints_from_form(request.form)
                    s["primary_category"]     = request.form.get("primary_category", "Sonstiges")
                    s["secondary_categories"] = request.form.getlist("secondary_categories")
                    s["relevant_roles"]       = request.form.getlist("relevant_roles")
                    s["ingestion_methods"]    = request.form.getlist("ingestion_methods")
                    s["priority"]             = request.form.get("priority", "mittel")
                    s["status"]               = request.form.get("status", "").strip()
                    s["zugang"]               = request.form.get("zugang", "").strip()
                    s["expected_update"]      = request.form.get("expected_update", "").strip()
                    s["notes"]                = request.form.get("notes", "").strip()
                    agent_hints = []
                    for i in ("1", "2"):
                        agent = request.form.get(f"agent_hint_{i}_agent", "").strip()
                        text  = request.form.get(f"agent_hint_{i}_text", "").strip()
                        if agent or text:
                            agent_hints.append({"agent": agent, "text": text})
                    s["agent_hints"] = agent_hints
                    break
            save_sources(sources_list)
            flash("Quelle aktualisiert.", "success")

        elif action == "delete":
            sid = request.form.get("source_id", "")
            save_sources([s for s in sources_list if s["id"] != sid])
            flash("Quelle gelöscht.", "success")

        elif action == "hints":
            cfg["source_prompt_hints"]      = request.form.get("source_prompt_hints", "").strip()
            cfg["source_output_format"]     = request.form.get("source_output_format", "").strip()
            save_settings(cfg)
            flash("Hinweise gespeichert.", "success")

        return redirect(url_for("sources"))

    default_output_fmt = (
        "**[KÜRZEL] Quelle:** Name der Quelle\n"
        "**Datum:** TT.MM.JJJJ (soweit bekannt)\n"
        "**Titel:** Kurzer beschreibender Titel\n"
        "**Zusammenfassung:** 2–4 Sätze, die den Inhalt und die Relevanz für Gesundheitseinrichtungen erklären\n"
        "**Betroffene Rollen:** (z.B. Praxisinhaber, Abrechnung, QM, Datenschutz)\n"
        "**Priorität:** Hoch / Mittel / Niedrig\n"
        "**Link:** Direkter Link zur Quelle\n\n---"
    )
    sources_list = load_sources()
    hints = cfg.get("source_prompt_hints", "")
    source_output_format = cfg.get("source_output_format", default_output_fmt)
    return render_template("sources.html",
        sources=sources_list, hints=hints,
        source_output_format=source_output_format,
        source_primary_categories=SOURCE_PRIMARY_CATEGORIES,
        source_secondary_categories=SOURCE_SECONDARY_CATEGORIES,
        source_relevant_roles=SOURCE_RELEVANT_ROLES,
        source_ingestion_methods=SOURCE_INGESTION_METHODS,
        source_zugang_options=SOURCE_ZUGANG_OPTIONS,
        source_update_options=SOURCE_UPDATE_OPTIONS,
        source_priorities=SOURCE_PRIORITIES,
        source_statuses=SOURCE_STATUSES,
        source_status_labels=SOURCE_STATUS_LABELS,
        region_options=REGION_OPTIONS)

@app.route("/admin/sources/import", methods=["POST"])
@role_required("admin", "superadmin")
def sources_import():
    rows = request.get_json(silent=True) or []
    if not rows:
        return jsonify({"ok": False, "msg": "Keine Daten"}), 400
    sources_list = load_sources()
    added = 0
    for row in rows[:20]:
        name = (row.get("name") or "").strip()
        if not name:
            continue
        sources_list.append({
            "id":                   str(uuid.uuid4()),
            "kuerzel":              (row.get("kuerzel") or "").strip(),
            "name":                 name,
            "url":                  (row.get("url") or "").strip(),
            "primary_category":     row.get("primary_category") or row.get("category") or "Sonstiges",
            "secondary_categories": row.get("secondary_categories") if isinstance(row.get("secondary_categories"), list) else [],
            "relevant_roles":       row.get("relevant_roles") if isinstance(row.get("relevant_roles"), list) else [],
            "priority":             row.get("priority") or "mittel",
            "status":               (row.get("status") or "").strip(),
            "zugang":               (row.get("zugang") or "").strip(),
            "expected_update":      (row.get("expected_update") or "").strip(),
            "notes":      (row.get("notes") or "").strip(),
            "created_at": datetime.now().strftime("%Y-%m-%d"),
        })
        added += 1
    save_sources(sources_list)
    return jsonify({"ok": True, "added": added})


# ── 403 handler ───────────────────────────────────────────────────────────────

@app.errorhandler(403)
def forbidden(e):
    return render_template("403.html"), 403


# ── Deploy webhook ────────────────────────────────────────────────────────────

@app.route("/deploy", methods=["POST"])
def deploy():
    import signal as _signal
    token = request.headers.get("X-Deploy-Token", "")
    if not DEPLOY_TOKEN or not hmac.compare_digest(token, DEPLOY_TOKEN):
        abort(403)
    repo = os.path.dirname(os.path.abspath(__file__))
    try:
        subprocess.run(["git", "pull", "--ff-only"], cwd=repo, check=True, capture_output=True)
        # Send SIGHUP to gunicorn master → graceful reload of all workers
        os.kill(os.getppid(), _signal.SIGHUP)
        return jsonify({"ok": True, "msg": "Pulled and reloading"})
    except subprocess.CalledProcessError as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True, port=5001)
